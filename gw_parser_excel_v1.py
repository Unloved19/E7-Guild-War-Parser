from __future__ import annotations
from pathlib import Path
from typing import Dict, List, Optional, Tuple
import time
import json
import threading
import tempfile
import shutil
try:
    import mss
    import numpy as np
except ImportError as exc:
    raise SystemExit("Missing dependency: mss or numpy. Install with: python -m pip install mss numpy") from exc


try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill
except ImportError as exc:
    raise SystemExit("Missing dependency: openpyxl. Install with: python -m pip install openpyxl") from exc

from gw_parser_core import *

# ============================================================================
# WORKBOOK CONFIGURATION
# ============================================================================
PROCESSING_SHEET = "Processing"

HEADER_FILL = "1F4E78"
SUBHEADER_FILL = "D9EAF7"

RAW_NAME_HEADER = "Raw Name"
MATCHED_NAME_HEADER = "Matched Workbook Name"

PROCESSING_RESULTS_HEADER = [
    "Screenshot", "Order", RAW_NAME_HEADER, MATCHED_NAME_HEADER,
    "Off W", "Off D", "Off L", "Def W", "Def D", "Def L", "Result String",
]

# ============================================================================
# SHARED DATA ACCESS HELPERS (2D list interface)
# ============================================================================

def _excel_to_2d_list(ws, max_col: int = 650) -> list:
    """Convert openpyxl worksheet to 2D list."""
    data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
        data.append([str(v) if v is not None else "" for v in row])
    return data


# ============================================================================
# WORKSHEET WRITING OPERATIONS
# ============================================================================

def ensure_processing_sheet(wb):
    if PROCESSING_SHEET in wb.sheetnames:
        del wb[PROCESSING_SHEET]
    return wb.create_sheet(PROCESSING_SHEET)


def style_header_row(ws, row: int, start_col: int, end_col: int, fill: str = HEADER_FILL):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row, col)
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(bold=True, color="FFFFFF")


def write_processing_sheet(
    ws, folder_name: str, sheet_name: str, extracted: List[OCRRow],
    normalized: List[Dict[str, object]], aliases_created_this_run: Dict[str, str],
) -> None:
    ws["A1"] = "Guild War Processing Output"
    ws["A2"] = "Folder"
    ws["B2"] = folder_name
    ws["A3"] = "Workbook Output Sheet"
    ws["B3"] = sheet_name
    style_header_row(ws, 1, 1, 4)
    style_header_row(ws, 2, 1, 2, SUBHEADER_FILL)
    style_header_row(ws, 3, 1, 2, SUBHEADER_FILL)

    start_row = 5
    ws.cell(start_row, 1, "Results Output")
    style_header_row(ws, start_row, 1, len(PROCESSING_RESULTS_HEADER))
    for idx, header in enumerate(PROCESSING_RESULTS_HEADER, start=1):
        ws.cell(start_row + 1, idx, header)
    style_header_row(ws, start_row + 1, 1, len(PROCESSING_RESULTS_HEADER), SUBHEADER_FILL)

    for i, row in enumerate(extracted, start=start_row + 2):
        values = [
            row.screenshot, row.order, row.raw_name, row.matched_name,
            row.off_w, row.off_d, row.off_l, row.def_w, row.def_d, row.def_l, row.result_string,
        ]
        for j, value in enumerate(values, start=1):
            ws.cell(i, j, value)

    norm_start = start_row + 4 + len(extracted)
    ws.cell(norm_start, 1, "Normalized Table")
    style_header_row(ws, norm_start, 1, len(PROCESSING_NORMALIZED_HEADER))
    for idx, header in enumerate(PROCESSING_NORMALIZED_HEADER, start=1):
        ws.cell(norm_start + 1, idx, header)
    style_header_row(ws, norm_start + 1, 1, len(PROCESSING_NORMALIZED_HEADER), SUBHEADER_FILL)
    for i, row in enumerate(normalized, start=norm_start + 2):
        for j, header in enumerate(PROCESSING_NORMALIZED_HEADER, start=1):
            ws.cell(i, j, row[header])

    preview_headers = [
        "Name", "Active flag", "Tokens Used", "Offense W", "Offense D", "Offense L",
        "Defense W", "Defense D", "Defense L",
    ]
    preview_start = norm_start + 4 + len(normalized)
    ws.cell(preview_start, 1, "Write Preview")
    style_header_row(ws, preview_start, 1, len(preview_headers))
    for idx, header in enumerate(preview_headers, start=1):
        ws.cell(preview_start + 1, idx, header)
    style_header_row(ws, preview_start + 1, 1, len(preview_headers), SUBHEADER_FILL)
    for i, row in enumerate(normalized, start=preview_start + 2):
        for j, header in enumerate(preview_headers, start=1):
            ws.cell(i, j, row[header])

    alias_start = preview_start + 4 + len(normalized)
    ws.cell(alias_start, 1, "New Aliases Created This Run")
    style_header_row(ws, alias_start, 1, 2)
    ws.cell(alias_start + 1, 1, "Raw OCR Name")
    ws.cell(alias_start + 1, 2, "Workbook Name")
    style_header_row(ws, alias_start + 1, 1, 2, SUBHEADER_FILL)
    row_idx = alias_start + 2
    for raw_name, workbook_name in sorted(aliases_created_this_run.items()):
        ws.cell(row_idx, 1, raw_name)
        ws.cell(row_idx, 2, workbook_name)
        row_idx += 1

    widths = {"A": 22, "B": 12, "C": 28, "D": 32, "E": 10, "F": 10, "G": 10, "H": 10, "I": 10, "J": 10, "K": 46}
    for col, width in widths.items():
        ws.column_dimensions[col].width = width



def clear_target_row_values(
    ws, target_row: int, section_maps: Dict[str, Dict[str, int]]
) -> None:
    for section_name in WRITE_SECTIONS:
        for col in section_maps[section_name].values():
            ws.cell(target_row, col).value = None


def write_to_rough_sheet(
    ws, target_row: int, normalized: List[Dict[str, object]],
    section_maps: Dict[str, Dict[str, int]],
) -> None:
    for row in normalized:
        name = row["Name"]
        mappings = {
            "Roster Active Status": row["Active flag"],
            "Tokens Tracking": row["Tokens Used"],
            "Offense Wins": row["Offense W"],
            "Offense Draws": row["Offense D"],
            "Offense Losses": row["Offense L"],
            "Defense Wins": row["Defense W"],
            "Defense Draws": row["Defense D"],
            "Defense Losses": row["Defense L"],
        }
        for section_name, value in mappings.items():
            col_map = section_maps[section_name]
            if name not in col_map:
                continue
            ws.cell(target_row, col_map[name], value)

# ============================================================================
# MAIN EXECUTION FLOW
# ============================================================================

def main():
    # --- DEBUG CLEANUP PROMPT ---
    if DEBUG_ROOT.exists():
        if ask_yes_no("Debug images from previous runs found. Delete them before starting?", default=True):
            cleanup_debug_images()

    # --- LOAD CONFIG ---
    config = load_config()
    workbook_path = BASE_DIR / config["excel"]["workbook_path"]
    sheet_name = config["excel"]["sheet_name"]
    override_col = config.get("override_col", "C")
    date_col = config.get("date_col", 15)

    print(f"Mode: Excel")
    print(f"Workbook: {workbook_path.name}")
    print(f"Sheet: {sheet_name}")

    run_mode = ask_run_mode()

    screenshots_folder, folder_name, video_temp_dir = prompt_and_get_screenshots()

    if not screenshots_folder.exists():
        raise FileNotFoundError(f"Could not find source: {screenshots_folder}")

    lock_file = workbook_path.parent / f"~${workbook_path.name}"
    if lock_file.exists():
        print(f"\nError: The workbook is currently open in Excel.")
        print(f"Detected lock file: {lock_file.name}")
        raise SystemExit(1)

    aliases_existing = load_aliases(ALIASES_PATH)
    aliases_working = dict(aliases_existing)
    aliases_created_this_run: Dict[str, str] = {}

    wb = load_workbook(workbook_path)
    if sheet_name not in wb.sheetnames:
        raise RuntimeError(f"Sheet missing: {sheet_name}")

    source_ws = wb[sheet_name]
    target_ws = wb[sheet_name]

    source_data = _excel_to_2d_list(source_ws)
    target_data = _excel_to_2d_list(target_ws)

    roster = read_roster(source_data)
    target_row = find_date_row(target_data, folder_name, date_col)
    target_section_maps = build_target_section_maps(target_data)
    overrides = read_column_a_overrides(target_data, roster, override_col)

    try:
        extracted, _contexts = extract_rows_from_folder(
            screenshots_folder, folder_name, roster,
            aliases_existing, aliases_working, aliases_created_this_run, overrides
        )
        normalized = build_normalized_table(roster, extracted)

        prev_active_status = get_previous_war_active_status(target_data, target_row, target_section_maps, date_col)
        resolve_active_inactive_status(normalized, extracted, prev_active_status)

        run_mode = handle_post_processing(run_mode, extracted, normalized, aliases_created_this_run)

        # Excel specific: Write the Processing tab regardless of dry/write
        proc_ws = ensure_processing_sheet(wb)
        write_processing_sheet(proc_ws, folder_name, sheet_name, extracted, normalized, aliases_created_this_run)

        if run_mode == "write":
            if row_has_existing_values(target_data, target_row, target_section_maps):
                if not ask_yes_no("Target row already has values in one or more write sections. Overwrite whole row?", default=False):
                    raise RuntimeError("Run cancelled by user at overwrite prompt.")

            clear_target_row_values(target_ws, target_row, target_section_maps)
            write_to_rough_sheet(target_ws, target_row, normalized, target_section_maps)

            if aliases_created_this_run:
                save_aliases(ALIASES_PATH, aliases_working)

            wb.save(workbook_path)

            print("\nDone.")
            print(f"Folder processed: {folder_name}")
            print(f"Workbook updated: {workbook_path.name}")
            print(f"Processing sheet rebuilt: {PROCESSING_SHEET}")
            print(f"Data written to: {sheet_name} (date row matched from folder name)")
            print(f"New aliases saved: {len(aliases_created_this_run)}")

    finally:
        if video_temp_dir and video_temp_dir.exists():
            shutil.rmtree(video_temp_dir)
            print("Cleaned up temporary capture files.")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
        raise SystemExit(1)
    except Exception as exc:
        print(f"\nError: {exc}")
        raise SystemExit(1)