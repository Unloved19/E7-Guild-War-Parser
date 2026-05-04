"""First-run setup wizard for the GW Parser.

Modes:
  - Non-interactive (called from GUI): python first_run_setup.py --mode excel --season "2026-1" ...
  - Partial reconfigure:               python first_run_setup.py --configure
  - Interactive CLI fallback:           python first_run_setup.py
"""

import re
import json
import sys
from pathlib import Path
from datetime import datetime

# ============================================================================
# PATHS & CONSTANTS
# ============================================================================

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
TOKEN_PATH = BASE_DIR / "token.json"
SECRET_PATH = BASE_DIR / "client_secret.json"

TEMPLATE_URL = "https://docs.google.com/spreadsheets/d/1gCnRQ0-K0Lvn5KjXmjpgWQZvp59P5oabsR0n1UEm8gk/edit?usp=sharing"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

RE_SHEET_ID = re.compile(r"/d/([a-zA-Z0-9_-]+)")


# ============================================================================
# CLI PROMPT HELPERS
# ============================================================================

def ask_yes_no(prompt: str, default: bool = False) -> bool:
    suffix = " [Y/n]: " if default else " [y/N]: "
    while True:
        answer = input(prompt + suffix).strip().lower()
        if answer == "":
            return default
        if answer in {"y", "yes"}:
            return True
        if answer in {"n", "no"}:
            return False
        print("Please answer y or n.")


def extract_spreadsheet_id(url: str) -> str:
    match = RE_SHEET_ID.search(url)
    if not match:
        raise ValueError("Could not find spreadsheet ID in that URL.")
    return match.group(1)


# ============================================================================
# GOOGLE SHEETS AUTHENTICATION
# ============================================================================

def run_auth() -> bool:
    """Runs the OAuth flow. Returns True on success."""
    try:
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
    except ImportError:
        print("Missing auth dependencies. Install with:")
        print("  python -m pip install google-auth-oauthlib google-auth-httplib2 google-api-python-client")
        return False

    if not SECRET_PATH.exists():
        print(f"\nMissing {SECRET_PATH.name}.")
        print("This file should have been included with the program.")
        print("If it's missing, contact the person who gave you this tool.")
        return False

    creds = None
    if TOKEN_PATH.exists():
        try:
            creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
        except ValueError:
            TOKEN_PATH.unlink()
            creds = None

    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            print("Refreshing expired token...", end=" ", flush=True)
            creds.refresh(Request())
            print("Done.")
        else:
            print("\nA browser window will open. Log into the Google account")
            print("that owns your Guild War spreadsheet and click 'Allow'.")
            input("Press ENTER to open the browser...")
            flow = InstalledAppFlow.from_client_secrets_file(str(SECRET_PATH), SCOPES)
            creds = flow.run_local_server(port=0)

        with open(TOKEN_PATH, "w") as f:
            json.dump(json.loads(creds.to_json()), f, indent=2)

    print("Authentication successful.")
    return True


# ============================================================================
# EXCEL: ROSTER & DATE WRITING
# ============================================================================

def write_roster_excel(roster: list, sheet_name: str, workbook_path: str = None) -> bool:
    """Write roster to Column B in Excel workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Error: openpyxl not installed.")
        return False

    wb_path = Path(workbook_path) if workbook_path else BASE_DIR / "GW Tracker.xlsx"
    lock_file = wb_path.parent / f"~${wb_path.name}"

    if lock_file.exists():
        print("Error: Workbook is open in Excel. Close it and try again.")
        return False
    if not wb_path.exists():
        print(f"Error: Workbook not found: {wb_path}")
        return False

    try:
        wb = load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
        else:
            ws = wb[sheet_name]

        for row in range(5, ws.max_row + 1):
            if ws.cell(row, 2).value is not None:
                ws.cell(row, 2).value = None

        for idx, name in enumerate(roster, start=5):
            ws.cell(idx, 2, name)

        wb.save(wb_path)
        print(f"Wrote {len(roster)} players to Column B in '{sheet_name}'.")
        return True
    except PermissionError:
        print("Error: Cannot save workbook. Is it open in Excel?")
        return False
    except Exception as exc:
        print(f"Error: {exc}")
        return False


def write_first_date_excel(sheet_name: str, date_str: str, workbook_path: str = None) -> bool:
    """Write first war date to O5 in Excel workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return False

    wb_path = Path(workbook_path) if workbook_path else BASE_DIR / "GW Tracker.xlsx"
    lock_file = wb_path.parent / f"~${wb_path.name}"

    if lock_file.exists() or not wb_path.exists():
        return False

    try:
        wb = load_workbook(wb_path)
        if sheet_name not in wb.sheetnames:
            return False
        ws = wb[sheet_name]
        parsed = datetime.strptime(date_str, "%Y%m%d")
        ws.cell(5, 15, parsed.date())
        wb.save(wb_path)
        print(f"Wrote first war date to O5: {parsed.strftime('%d/%m/%Y')}")
        return True
    except Exception:
        return False


# ============================================================================
# GOOGLE SHEETS: ROSTER & DATE WRITING
# ============================================================================

def write_roster_gs(roster: list, spreadsheet_id: str, sheet_name: str) -> bool:
    """Write roster to Column B in Google Sheet."""
    try:
        import gspread
        from google.oauth2.credentials import Credentials
        from google_auth.transport.requests import Request
    except ImportError:
        print("Error: gspread not installed.")
        return False

    try:
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
        if not creds.valid:
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                print("Error: Auth token invalid. Run setup again.")
                return False

        client = gspread.authorize(creds)
        ws = client.open_by_key(spreadsheet_id).worksheet(sheet_name)

        data = ws.get_all_values()
        updates = []

        for row in range(5, len(data) + 5):
            if row <= len(data) and len(data[row - 1]) >= 2 and data[row - 1][1]:
                updates.append({"range": f"B{row}", "values": [[""]]})

        for idx, name in enumerate(roster, start=5):
            updates.append({"range": f"B{idx}", "values": [[name]]})

        if updates:
            ws.batch_update(updates)

        print(f"Wrote {len(roster)} players to Column B in '{sheet_name}'.")
        return True
    except Exception as exc:
        print(f"Error: {exc}")
        return False


def write_first_date_gs(spreadsheet_id: str, sheet_name: str, date_str: str) -> bool:
    """Write first war date to O5 in Google Sheet."""
    try:
        import gspread
        from google.oauth2.credentials import Credentials
        from google_auth_transport.requests import Request
    except ImportError:
        return False

    try:
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
        if not creds.valid and creds.expired and creds.refresh_token:
            creds.refresh(Request())

        client = gspread.authorize(creds)
        ws = client.open_by_key(spreadsheet_id).worksheet(sheet_name)
        formatted = datetime.strptime(date_str, "%Y%m%d").strftime("%d/%m/%Y")

        ws.batch_update(
            [{"range": "O5", "values": [[formatted]]}],
            value_input_option="USER_ENTERED",
        )
        print(f"Wrote first war date to O5: {formatted}")
        return True
    except Exception:
        return False


# ============================================================================
# NON-INTERACTIVE SETUP (called from GUI)
# ============================================================================

def setup_non_interactive(mode, season, spreadsheet_id=None, first_date=None, roster=None):
    """Run setup with provided arguments (no prompts)."""
    config = {
        "mode": mode,
        "override_col": "C",
        "date_col": 15,
        "guild_name": None,  # Auto-detect from filename/title
    }

    if mode == "google_sheet":
        if not spreadsheet_id:
            print("Error: --spreadsheet-id required for Google Sheets mode.")
            return False
        config["google_sheet"] = {
            "spreadsheet_id": spreadsheet_id,
            "source_sheet": season,
            "target_sheet": season,
        }
    else:
        config["excel"] = {
            "workbook_path": "GW Tracker.xlsx",
            "sheet_name": season,
        }

    if roster:
        if mode == "google_sheet":
            if not write_roster_gs(roster, spreadsheet_id, season):
                return False
        else:
            if not write_roster_excel(roster, season):
                return False

    if first_date:
        if mode == "google_sheet":
            write_first_date_gs(spreadsheet_id, season, first_date)
        else:
            write_first_date_excel(season, first_date)

    CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")
    print(f"Config saved to: {CONFIG_PATH.name}")
    return True

def change_spreadsheet_non_interactive(new_spreadsheet_id: str) -> bool:
    if not CONFIG_PATH.exists():
        print("Error: No existing config found.")
        return False

    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    mode = config.get("mode", "google_sheet")

    if mode == "google_sheet":
        config["google_sheet"]["spreadsheet_id"] = new_spreadsheet_id
        # Preserve guild_name (don't reset it)
        if "guild_name" not in config:
            config["guild_name"] = None
        CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")
        print(f"Spreadsheet ID updated to: {new_spreadsheet_id}")
        return True
    else:
        print("Error: Not in Google Sheets mode.")
        return False

# ============================================================================
# PARTIAL RECONFIGURE (called from GUI Configure button)
# ============================================================================

def configure_partial():
    """Reconfigure spreadsheet/workbook only, keep mode and season."""
    if not CONFIG_PATH.exists():
        print("No existing config found. Run full setup first.")
        return False

    config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
    mode = config.get("mode", "google_sheet")

    if mode == "google_sheet":
        current_id = config["google_sheet"].get("spreadsheet_id", "")
        print(f"Current spreadsheet ID: {current_id}")
        if not ask_yes_no("Change to a different spreadsheet?", default=False):
            return True

        while True:
            url = input("  Paste new Google Sheet URL: ").strip()
            try:
                new_id = extract_spreadsheet_id(url)
                break
            except ValueError:
                print("  Invalid URL.")

        config["google_sheet"]["spreadsheet_id"] = new_id

        if ask_yes_no("Re-run authentication?", default=False):
            if not run_auth():
                return False
    else:
        current_path = config["excel"].get("workbook_path", "GW Tracker.xlsx")
        print(f"Current workbook: {current_path}")
        if not ask_yes_no("Change to a different workbook?", default=False):
            return True

        new_path = input("  New workbook path (or filename): ").strip()
        config["excel"]["workbook_path"] = new_path or "GW Tracker.xlsx"

    CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")
    print(f"Config saved to: {CONFIG_PATH.name}")
    return True


# ============================================================================
# INTERACTIVE CLI SETUP (fallback / direct run)
# ============================================================================

def collect_config_gs(existing: dict = None) -> dict:
    """Interactive Google Sheets config collection."""
    print("=" * 50)
    print("GW Parser - Google Sheets Setup")
    print("=" * 50)

    if existing and existing.get("google_sheet", {}).get("spreadsheet_id"):
        print(f"\nCurrent spreadsheet ID: {existing['google_sheet']['spreadsheet_id']}")
        if not ask_yes_no("Change to a different spreadsheet?", default=False):
            spreadsheet_id = existing["google_sheet"]["spreadsheet_id"]
        else:
            spreadsheet_id = None
    else:
        spreadsheet_id = None

    if not spreadsheet_id:
        print("\nStep 1: Get your Google Sheet")
        print(f"  1. Open this link: {TEMPLATE_URL}")
        print("  2. Click 'Make a Copy' to create your own copy")
        print("  3. Copy the URL of your new spreadsheet")
        while True:
            url = input("  Paste your Google Sheet URL here: ").strip()
            try:
                spreadsheet_id = extract_spreadsheet_id(url)
                print(f"  Extracted ID: {spreadsheet_id}")
                break
            except ValueError:
                print("  Invalid URL.")

    default_sheet = existing.get("google_sheet", {}).get("source_sheet", "GW Tracker") if existing else "GW Tracker"
    sheet_name = input(f"\nSheet name [default: {default_sheet}]: ").strip() or default_sheet

    print(f"\nVerifying access to '{sheet_name}'...")
    try:
        import gspread
        from google.oauth2.credentials import Credentials
        from google.auth.transport.requests import Request

        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
        if not creds.valid:
            if creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                raise SystemExit("Auth token invalid.")

        client = gspread.authorize(creds)
        sheet_names = [ws.title for ws in client.open_by_key(spreadsheet_id).worksheets()]
        if sheet_name not in sheet_names:
            print(f"  Available: {', '.join(sheet_names)}")
            print(f"  '{sheet_name}' not found!")
            if not ask_yes_no("Continue anyway?", default=False):
                raise SystemExit("Setup cancelled.")
        print("  Connected successfully.")
    except SystemExit:
        raise
    except Exception as exc:
        print(f"  Warning: {exc}")
        if not ask_yes_no("Continue anyway?", default=False):
            raise SystemExit("Setup cancelled.")

    return {
        "spreadsheet_id": spreadsheet_id,
        "source_sheet": sheet_name,
        "target_sheet": sheet_name,
    }


def print_post_setup_instructions(config: dict) -> None:
    print("\n" + "=" * 50)
    print("Setup complete! Before running the parser:")
    print("=" * 50)
    print("  1. Open your spreadsheet/workbook")
    print("  2. Add guild member names in Column B (starting at row 5)")
    print("  3. Enter the first war date in Column O (row 5)")
    print("  4. Run the parser to record your first war")
    print("=" * 50)


def interactive_setup():
    """Full interactive CLI setup."""
    if CONFIG_PATH.exists():
        existing = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        print("Found existing configuration.")
        if not ask_yes_no("Reconfigure?", default=False):
            return
    else:
        existing = None

    # Mode selection
    if existing and existing.get("mode"):
        mode = existing["mode"]
        print(f"\nCurrent mode: {mode}")
        if ask_yes_no("Change mode?", default=False):
            mode = "google_sheet" if input("1=GS, 2=Excel: ").strip() == "1" else "excel"
    else:
        mode = "google_sheet" if input("1=Google Sheets, 2=Excel: ").strip() == "1" else "excel"

    # Mode-specific setup
    if mode == "google_sheet":
        if not TOKEN_PATH.exists():
            print("\nNo authentication found.")
            if not run_auth():
                raise SystemExit(1)
        gs_config = collect_config_gs(existing)
        config = {"mode": mode, "google_sheet": gs_config, "override_col": "C", "date_col": 15, "guild_name": None}
        season = gs_config["source_sheet"]
        spreadsheet_id = gs_config["spreadsheet_id"]
    else:
        season = input("\nSeason name (e.g. 2026-1): ").strip() or "2026-1"
        workbook = input(f"Workbook path [default: GW Tracker.xlsx]: ").strip() or "GW Tracker.xlsx"
        config = {"mode": mode, "excel": {"workbook_path": workbook, "sheet_name": season}, "override_col": "C", "date_col": 15, "guild_name": None}
        spreadsheet_id = None

    # Roster entry
    print("\nPaste player names, one per line (empty line to finish):")
    roster = []
    while True:
        name = input().strip()
        if not name:
            break
        roster.append(name)

    if roster:
        print(f"\nWriting {len(roster)} players to Column B...")
        if mode == "google_sheet":
            write_roster_gs(roster, spreadsheet_id, season)
        else:
            write_roster_excel(roster, season, config["excel"]["workbook_path"])

    # First date
    date_input = input("\nFirst war date (YYYYMMDD, or blank to skip): ").strip()
    if date_input and len(date_input) == 8 and date_input.isdigit():
        if mode == "google_sheet":
            write_first_date_gs(spreadsheet_id, season, date_input)
        else:
            write_first_date_excel(season, date_input, config["excel"]["workbook_path"])

    CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")
    print(f"\nConfig saved to: {CONFIG_PATH.name}")
    print_post_setup_instructions(config)


# ============================================================================
# ENTRY POINT
# ============================================================================

def main():
    args = sys.argv[1:]

    # Non-interactive mode (called from GUI)
    if "--mode" in args:
        mode = args[args.index("--mode") + 1] if args.index("--mode") + 1 < len(args) else None
        if mode not in ("excel", "google_sheet"):
            raise SystemExit("Invalid --mode. Use 'excel' or 'google_sheet'.")

        season = args[args.index("--season") + 1] if "--season" in args and args.index("--season") + 1 < len(args) else "2026-1"
        spreadsheet_id = args[args.index("--spreadsheet-id") + 1] if "--spreadsheet-id" in args and args.index("--spreadsheet-id") + 1 < len(args) else None
        first_date = args[args.index("--first-date") + 1] if "--first-date" in args and args.index("--first-date") + 1 < len(args) else None

        # Remaining args after known flags are roster names
        known_values = {"excel", "google_sheet", season, spreadsheet_id, first_date, ""}
        roster = [a for a in args if not a.startswith("--") and a not in known_values]

        success = setup_non_interactive(mode, season, spreadsheet_id, first_date, roster)
        raise SystemExit(0 if success else 1)

    # Partial reconfigure mode
    if "--configure" in args:
        success = configure_partial()
        raise SystemExit(0 if success else 1)
    
    # Change spreadsheet non-interactively (called from GUI)
    if "--set-spreadsheet-id" in args:
        new_id = args[args.index("--set-spreadsheet-id") + 1] if args.index("--set-spreadsheet-id") + 1 < len(args) else None
        if not new_id:
            raise SystemExit("Error: --set-spreadsheet-id requires a value.")
        success = change_spreadsheet_non_interactive(new_id)
        raise SystemExit(0 if success else 1)

    # Interactive CLI mode
    interactive_setup()


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\nCancelled.")
    except SystemExit:
        raise
    except Exception as exc:
        print(f"\nSetup error: {exc}")
        raise SystemExit(1)