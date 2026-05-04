"""Exports a war record summary from Google Sheets to a trimmed PNG.

Can be run standalone:
    python war_record_exporter.py 20260501

Or called from the main parser after a successful write.
"""

import sys
import re
import json
from datetime import datetime
from pathlib import Path
from typing import Optional, List

import requests
import fitz  # PyMuPDF
from PIL import Image, ImageChops, ImageOps
import gspread
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request

from openpyxl import load_workbook

# ============================================================================
# PATHS & CONFIG
# ============================================================================

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
TOKEN_PATH = BASE_DIR / "token.json"
SECRET_PATH = BASE_DIR / "client_secret.json"
WAR_RECORDS_DIR = BASE_DIR / "War Records"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

SEASON_TOTAL_TOKENS = 108
WAR_TOKEN_STEP = 3
EXPORT_RANGE = "B2:S45"
DATE_COL = 15  # Column O


# ============================================================================
# SHARED HELPERS
# ============================================================================

def load_config() -> dict:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"Missing {CONFIG_PATH.name}")
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def get_gspread_client() -> gspread.Client:
    if not SECRET_PATH.exists():
        raise SystemExit(f"Missing {SECRET_PATH.name}. Run setup_auth.py first.")
    creds = None
    if TOKEN_PATH.exists():
        creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            raise SystemExit("Token expired or missing. Run setup_auth.py again.")
    return gspread.authorize(creds)

def get_auth_token() -> str:
    """Reads the OAuth token directly from file, bypassing gspread internals."""
    if not TOKEN_PATH.exists():
        raise SystemExit("Token expired or missing. Run setup_auth.py again.")
    creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
    if not creds or not creds.valid:
        if creds and creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            raise SystemExit("Token expired or missing. Run setup_auth.py again.")
    return creds.token


def cell_val(data: List[List], row: int, col: int) -> Optional[str]:
    if row < 1 or row > len(data):
        return None
    row_data = data[row - 1]
    if col < 1 or col > len(row_data):
        return None
    val = row_data[col - 1]
    return val if val != "" else None


def sheet_date_to_folder_date_str(value) -> Optional[str]:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%Y/%m/%d", "%Y%m%d"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y%m%d")
            except ValueError:
                continue
    return None

def get_guild_name(config: dict, spreadsheet_title: str = None, workbook_path: Path = None) -> str:
    """Get guild name from config, spreadsheet title, or workbook filename."""
    # 1. Explicit config value
    if config.get("guild_name"):
        return config["guild_name"]
    
    # 2. Auto-detect from spreadsheet title (Google Sheets)
    if spreadsheet_title:
        return spreadsheet_title
    
    # 3. Auto-detect from workbook filename (Excel)
    if workbook_path:
        stem = workbook_path.stem  # e.g., "Zodiac GW Tracker"
        # Remove common suffixes
        for suffix in [" GW Tracker", " Tracker", " GW"]:
            if stem.endswith(suffix):
                stem = stem[:-len(suffix)]
                break
        return stem.strip() if stem.strip() else "Guild"
    
    return "Guild"

# ============================================================================
# IMAGE PROCESSING
# ============================================================================

def trim_white_margins(img: Image.Image, padding: int = 8) -> Image.Image:
    rgb = img.convert("RGB")
    bg = Image.new("RGB", rgb.size, (255, 255, 255))
    diff = ImageChops.difference(rgb, bg)
    bbox = diff.getbbox()
    if bbox is None:
        return rgb
    left, top, right, bottom = bbox
    cropped = rgb.crop((left, top, right, bottom))
    return ImageOps.expand(cropped, border=padding, fill="white")


# ============================================================================
# TOKEN CALCULATION
# ============================================================================

def read_war_dates(data: List[List]) -> List[str]:
    """Reads all war dates from column O, returns sorted unique YYYYMMDD list."""
    dates = []
    for row in range(5, len(data) + 1):
        val = cell_val(data, row, DATE_COL)
        if val:
            parsed = sheet_date_to_folder_date_str(val)
            if parsed:
                dates.append(parsed)
    return sorted(set(dates))


def calculate_used_tokens(war_dates: List[str], current_date: str) -> int:
    """Calculates tokens used based on war date position in the season."""
    if current_date not in war_dates:
        first = war_dates[0] if war_dates else "none"
        raise RuntimeError(
            f"Date {current_date} not found in sheet dates "
            f"(starts with {first})."
        )
    war_number = war_dates.index(current_date) + 1  # 1-indexed
    return war_number * WAR_TOKEN_STEP


# ============================================================================
# EXPORT LOGIC
# ============================================================================

def export_to_png_excel(
    folder_name: str,
    config: dict,
    dpi: int = 300,
) -> Path:
    """Export Excel range to PDF via win32com, then convert to PNG with PyMuPDF."""
    try:
        import win32com.client as win32  # type: ignore
    except ImportError:
        raise SystemExit(
            "win32com is not available.\n"
            "This feature requires Windows with Microsoft Excel installed.\n"
            "Use Google Sheets mode instead, or install Excel."
        )

    wb_path = BASE_DIR / config["excel"]["workbook_path"]
    sheet_name = config["excel"]["sheet_name"]
    date_col = config.get("date_col", 15)
    export_range = "B2:S45"

    # Get guild name from workbook filename
    guild_name = get_guild_name(config, workbook_path=wb_path)

    # Read data for token calculation
    wb_read = load_workbook(wb_path, data_only=True, read_only=True)
    ws_read = wb_read[sheet_name]
    data = _excel_to_2d_list(ws_read, max_col=650)
    wb_read.close()

    war_dates = read_war_dates(data, date_col)
    used = calculate_used_tokens(war_dates, folder_name)

    print(f"Exporting {export_range} from {sheet_name} via Excel...", end=" ", flush=True)

    # Export to temporary PDF via Excel's native PDF export
    temp_pdf = BASE_DIR / f"__tmp_export_{folder_name}.pdf"

    excel = win32.Dispatch("Excel.Application")
    excel.Visible = False
    excel.DisplayAlerts = False
    try:
        wb = excel.Workbooks.Open(str(wb_path))
        ws = wb.Sheets(sheet_name)
        ws.Range(export_range).Select()

        ws.ExportAsFixedFormat(
            0,              # Type: xlTypePDF
            str(temp_pdf),  # Filename: MUST be 2nd parameter
            0,              # Quality: xlQualityStandard  
            0,              # IncludeDocProperties
            0,              # IgnorePrintAreas
            0,              # From
            0,              # To
            0,              # OpenAfterPublish
        )

        wb.Close(False)
    finally:
        excel.Quit()

    # Convert PDF to PNG using PyMuPDF
    doc = fitz.open(str(temp_pdf))
    page = doc.load_page(0)
    scale = dpi / 72.0
    matrix = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=matrix, alpha=False)

    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    img = trim_white_margins(img, padding=8)

    WAR_RECORDS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"{guild_name} {used}-{SEASON_TOTAL_TOKENS} {folder_name}.png"
    out_path = WAR_RECORDS_DIR / filename
    img.save(str(out_path), "PNG")
    print(f"Saved: {filename}")

    # Cleanup temp PDF
    if temp_pdf.exists():
        temp_pdf.unlink()

    return out_path


def _excel_to_2d_list(ws, max_col: int = 650) -> list:
    """Converts openpyxl worksheet to 2D list (same as sync_roster)."""
    data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
        data.append([str(v) if v is not None else "" for v in row])
    return data


def export_to_png(
    client: gspread.Client,
    spreadsheet_id: str,
    sheet_name: str,
    folder_name: str,
    dpi: int = 300,
) -> Path:
    config = load_config()
    ws = client.open_by_key(spreadsheet_id).worksheet(sheet_name)
    gid = ws.id

    # Get guild name from spreadsheet title or config
    guild_name = get_guild_name(config, spreadsheet_title=ws.spreadsheet.title)

    data = ws.get_all_values()
    war_dates = read_war_dates(data)
    used = calculate_used_tokens(war_dates, folder_name)

    url = f"https://docs.google.com/spreadsheets/d/{spreadsheet_id}/export"
    params = {
        "format": "pdf",
        "gid": gid,
        "range": EXPORT_RANGE,
    }
    headers = {"Authorization": f"Bearer {get_auth_token()}"}

    print(f"Exporting {EXPORT_RANGE} from {sheet_name}...", end=" ", flush=True)
    response = requests.get(url, headers=headers, params=params)
    response.raise_for_status()
    print("Downloaded.", end=" ", flush=True)

    doc = fitz.open(stream=response.content, filetype="pdf")
    page = doc.load_page(0)
    scale = dpi / 72.0
    matrix = fitz.Matrix(scale, scale)
    pix = page.get_pixmap(matrix=matrix, alpha=False)

    img = Image.frombytes("RGB", (pix.width, pix.height), pix.samples)
    img = trim_white_margins(img, padding=8)

    WAR_RECORDS_DIR.mkdir(parents=True, exist_ok=True)
    filename = f"{guild_name} {used}-{SEASON_TOTAL_TOKENS} {folder_name}.png"
    out_path = WAR_RECORDS_DIR / filename

    img.save(str(out_path), "PNG")
    print(f"Saved: {filename}")
    return out_path


# ============================================================================
# ENTRY POINT
# ============================================================================

def main():
    if len(sys.argv) < 2:
        folder_name = input("Enter war date (YYYYMMDD): ").strip()
    else:
        folder_name = sys.argv[1]

    if not re.fullmatch(r"\d{8}", folder_name):
        raise SystemExit("Date must be 8 digits in YYYYMMDD format.")

    config = load_config()
    backend_mode = config.get("mode", "google_sheet")

    if backend_mode == "excel":
        export_to_png_excel(folder_name, config)
    else:
        gs_config = config.get("google_sheet", config)
        client = get_gspread_client()
        export_to_png(
            client,
            gs_config.get("spreadsheet_id", config.get("spreadsheet_id")),
            gs_config.get("target_sheet", config.get("target_sheet")),
            folder_name
        )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\nError: {exc}")
        raise SystemExit(1)