"""GW Parser GUI — Terminal-in-a-window."""

# ============================================================================
# IMPORTS
# ============================================================================

import sys
import os
import io
import re
import json
import importlib.util
import threading
import queue
import subprocess
import tempfile
import shutil
import time
from pathlib import Path
import mss
import numpy as np
import cv2
import easyocr

# ============================================================================
# TCL/TK AUTO-FIX - Must run before importing tkinter
# ============================================================================

def _fix_tcl_library():
    """Attempt to locate and set TCL_LIBRARY if tkinter will fail."""
    candidate_paths = [
        Path(os.environ.get("CONDA_PREFIX", "")) / "tcl" / "tcl8.6" if "CONDA_PREFIX" in os.environ else None,
        Path.home() / "anaconda3" / "tcl" / "tcl8.6",
        Path.home() / "anaconda3" / "envs" / "egb439" / "tcl" / "tcl8.6",
        Path(sys.executable).parent / "tcl" / "tcl8.6",
        Path(sys.executable).parent.parent / "tcl" / "tcl8.6",
    ]
    candidate_paths = [p for p in candidate_paths if p is not None]
    
    for tcl_path in candidate_paths:
        init_tcl = tcl_path / "init.tcl"
        if init_tcl.exists():
            os.environ["TCL_LIBRARY"] = str(tcl_path)
            tk_path = tcl_path.parent / "tk8.6"
            if tk_path.exists():
                os.environ["TK_LIBRARY"] = str(tk_path)
            print(f"[Tcl Fix] Found Tcl at: {tcl_path}")
            return True
    return False

_tcl_fixed = _fix_tcl_library()

try:
    import customtkinter as ctk
except ImportError:
    raise SystemExit(
        "Missing customtkinter.\nInstall with: python -m pip install customtkinter"
    )

try:
    import tkinter as tk
    _test_tk = tk.Tk()
    _test_tk.destroy()
except Exception as e:
    print("\n" + "=" * 60)
    print("ERROR: Tcl/Tk is not properly installed for this Python.")
    print("=" * 60)
    print(f"\nDetails: {e}")
    print("\nSolutions:")
    print("  1. Install Python 3.11 or 3.12 from python.org")
    print("     (Check 'tcl/tk and IDLE' during installation)")
    print("  2. Set environment variables manually:")
    print("     set TCL_LIBRARY=path\\to\\tcl8.6")
    print("     set TK_LIBRARY=path\\to\\tk8.6")
    print("  3. Use a conda environment with tk installed:")
    print("     conda install -c conda-forge tk")
    print("=" * 60 + "\n")
    raise SystemExit(1)

# ============================================================================
# CONSTANTS & PATHS
# ============================================================================

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
SCREENSHOTS_ROOT = BASE_DIR / "War Screenshots"

# ============================================================================
# STDIO REDIRECTORS
# ============================================================================

class StdoutRedirector(io.TextIOBase):
    """Redirects stdout to a thread-safe queue for GUI console display."""
    
    def __init__(self):
        self.queue = queue.Queue()

    def write(self, text: str) -> int:
        if text:
            self.queue.put(text)
        return len(text) if text else 0

    def flush(self) -> None:
        pass


class StdinRedirector:
    """Redirects stdin to a thread-safe queue for GUI input collection."""
    
    def __init__(self):
        self.queue = queue.Queue()

    def readline(self):
        return self.queue.get() + "\n"


# ============================================================================
# DIALOG WINDOWS
# ============================================================================

class DateDialog(ctk.CTkToplevel):
    """Modal dialog for entering a war date (YYYYMMDD)."""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Record War")
        self.geometry("360x155")
        self.result = None
        self.grab_set()

        ctk.CTkLabel(
            self,
            text="Enter war date (YYYYMMDD):",
            font=("Segoe UI", 14),
        ).pack(pady=(25, 10))

        self.entry = ctk.CTkEntry(
            self,
            font=("Segoe UI", 14),
            justify="center",
            width=280,
        )
        self.entry.pack(pady=(0, 5))

        ctk.CTkButton(
            self,
            text="Start",
            command=self._submit,
            height=42,
            font=("Segoe UI", 13),
        ).pack(pady=(5, 10))

    def _submit(self):
        val = self.entry.get().strip()
        if len(val) != 8 or not val.isdigit():
            return
        self.result = val
        self.grab_release()
        self.destroy()

    def destroy(self):
        self.grab_release()
        super().destroy()


class SourceChoiceDialog(ctk.CTkToplevel):
    """Modal dialog for choosing screenshot source mode."""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Data Source")
        self.geometry("320x160")
        self.result = None
        self.grab_set()

        ctk.CTkLabel(
            self,
            text="Select data source:",
            font=("Segoe UI", 14),
        ).pack(pady=(25, 15))

        ctk.CTkButton(
            self,
            text="Screenshot Folder",
            width=250,
            height=48,
            font=("Segoe UI", 13),
            command=lambda: self._choose("folder"),
        ).pack(pady=(0, 8))

        ctk.CTkButton(
            self,
            text="Live Screen Capture",
            width=250,
            height=48,
            font=("Segoe UI", 13),
            command=lambda: self._choose("live"),
        ).pack(pady=(0, 10))

    def _choose(self, choice):
        self.result = choice
        self.grab_release()
        self.destroy()

class NewSeasonDialog(ctk.CTkToplevel):
    """Modal dialog for new season setup."""
    
    def __init__(self, parent, current_season: str):
        super().__init__(parent)
        self.title("New Season")
        self.geometry("380x220")
        self.result = None
        self.grab_set()

        ctk.CTkLabel(
            self,
            text="New Season Setup",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(20, 15))

        # Auto-increment season name
        parts = current_season.rsplit("-", 1)
        if len(parts) == 2 and parts[1].strip().isdigit():
            default_name = f"{parts[0]}-{int(parts[1].strip()) + 1}"
        else:
            default_name = f"{current_season} (2)"

        ctk.CTkLabel(self, text="Season name:", font=("Segoe UI", 12)).pack(anchor="w", padx=(40, 0))
        self.name_entry = ctk.CTkEntry(self, placeholder_text=default_name, width=300, font=("Segoe UI", 13))
        self.name_entry.pack(padx=(40, 0), pady=(2, 10))

        ctk.CTkLabel(self, text="First war date (YYYYMMDD):", font=("Segoe UI", 12)).pack(anchor="w", padx=(40, 0))
        self.date_entry = ctk.CTkEntry(self, width=300, font=("Segoe UI", 13), justify="center")
        self.date_entry.pack(padx=(40, 0), pady=(2, 15))
        self.date_entry.focus_set()
        self.date_entry.bind("<Return>", lambda e: self._submit())

        ctk.CTkButton(self, text="Create", width=200, command=self._submit).pack(pady=(5, 10))

    def _submit(self):
        season = self.name_entry.get().strip()
        date_val = self.date_entry.get().strip()
        if not date_val or len(date_val) != 8 or not date_val.isdigit():
            return
        self.result = {"season_name": season if season else None, "first_date": date_val}
        self.grab_release()
        self.destroy()

class ConfigureDialog(ctk.CTkToplevel):
    """Modal dialog for configure options."""
    
    def __init__(self, parent):
        super().__init__(parent)
        self.title("Configure")
        self.geometry("320x200")
        self.result = None
        self.grab_set()

        ctk.CTkLabel(
            self,
            text="Configure Options",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(25, 20))

        ctk.CTkButton(
            self,
            text="Change spreadsheet/workbook",
            width=250,
            height=48,
            font=("Segoe UI", 13),
            command=lambda: self._choose("1"),
        ).pack(pady=(0, 8))

        ctk.CTkButton(
            self,
            text="Re-run full setup wizard",
            width=250,
            height=48,
            font=("Segoe UI", 13),
            command=lambda: self._choose("2"),
        ).pack(pady=(0, 10))

    def _choose(self, choice):
        self.result = choice
        self.grab_release()
        self.destroy()

class ChangeSpreadsheetDialog(ctk.CTkToplevel):
    """Modal dialog for entering new spreadsheet URL."""
    
    def __init__(self, parent, current_id: str):
        super().__init__(parent)
        self.title("Change Spreadsheet")
        self.geometry("500x160")
        self.result = None
        self.grab_set()

        ctk.CTkLabel(
            self,
            text="Change Spreadsheet",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(20, 5))

        ctk.CTkLabel(
            self,
            text=f"Current: {current_id}",
            font=("Segoe UI", 10),
            text_color="gray50",
        ).pack()

        self.entry = ctk.CTkEntry(
            self,
            placeholder_text="https://docs.google.com/spreadsheets/d/...",
            font=("Segoe UI", 12),
            width=460,
        )
        self.entry.pack(pady=(10, 5))
        self.entry.focus_set()
        self.entry.bind("<Return>", lambda e: self._submit())

        ctk.CTkButton(
            self,
            text="Update",
            width=200,
            command=self._submit,
        ).pack(pady=(5, 10))

    def _submit(self):
        url = self.entry.get().strip()
        if not url:
            return
        match = re.search(r"/d/([a-zA-Z0-9_-]+)", url)
        if not match:
            return
        self.result = match.group(1)
        self.grab_release()
        self.destroy()
        
# ============================================================================
# FIRST RUN WIZARD
# ============================================================================

class FirstRunWizard(ctk.CTkToplevel):
    """Multi-step wizard for initial application setup."""
    
    # Screen indices
    SCREEN_MODE = 0
    SCREEN_SEASON = 1
    SCREEN_LINK_OR_PATH = 2
    SCREEN_DATE = 3
    SCREEN_ROSTER = 4
    SCREEN_AUTH = 5
    SCREEN_DONE = 6
    SCREEN_COUNT = 7

    def __init__(self, parent):
        super().__init__(parent)
        self.title("First Run Setup")
        self.geometry("500x420")
        self.result = None
        self.grab_set()

        # Main container
        container = ctk.CTkFrame(self, fg_color="transparent")
        container.pack(fill="both", expand=True, padx=30, pady=20)

        # Title
        ctk.CTkLabel(
            container,
            text="Guild War Parser - First Run Setup",
            font=("Segoe UI", 20, "bold"),
        ).pack(pady=(0, 25))

        # Content frame (swapped between screens)
        self.content = ctk.CTkFrame(container, fg_color="transparent")
        self.content.pack(fill="both", expand=True)

        # Navigation frame
        nav = ctk.CTkFrame(container, fg_color="transparent")
        nav.pack(fill="x", side="bottom", pady=(15, 0))

        self.btn_back = ctk.CTkButton(
            nav, text="Back", width=120,
            command=self._go_back,
        )
        self.btn_back.pack(side="left", padx=(0, 10))

        self.btn_next = ctk.CTkButton(
            nav, text="Next →", width=120,
            command=self._go_next,
        )
        self.btn_next.pack(side="right", padx=(10, 0))

        # State
        self._screen = 0
        self._data = {}

        # Entry references (set by screen builders)
        self._season_entry = None
        self._link_entry = None
        self._date_entry = None
        self._roster_textbox = None

        # Show first screen
        self._show_current_screen()

    # -------------------------------------------------------------------------
    # Navigation
    # -------------------------------------------------------------------------

    def _clear_content(self):
        """Destroy all widgets in the content frame."""
        for widget in self.content.winfo_children():
            widget.destroy()

    def _go_back(self):
        """Navigate to the previous screen."""
        if self._screen > 0:
            self._screen -= 1
            self._show_current_screen()

    def _go_next(self):
        if not self._validate_current_screen():
            return

        # Capture mode when leaving mode screen
        if self._screen == self.SCREEN_MODE:
            self._data["mode"] = self._mode_var.get()

        if self._screen < self.SCREEN_DONE:
            self._screen += 1
            # Skip auth screen for Excel mode
            if self._screen == self.SCREEN_AUTH:
                if self._data.get("mode") == "excel":
                    self._screen = self.SCREEN_DONE
            self._show_current_screen()
        else:
            # On "Done" screen, finish the wizard
            self._finish()

    def _validate_current_screen(self) -> bool:
        """Validate the current screen's input. Returns True if valid."""
        if self._screen == self.SCREEN_SEASON:
            val = self._season_entry.get().strip()
            if not val:
                return False
            self._data["season_name"] = val
            return True

        elif self._screen == self.SCREEN_LINK_OR_PATH:
            mode = self._data.get("mode", "google_sheet")
            if mode == "google_sheet":
                link = self._link_entry.get().strip()
                if not link:
                    return False
                match = re.search(r"/d/([a-zA-Z0-9_-]+)", link)
                if not match:
                    return False
                self._data["spreadsheet_id"] = match.group(1)
            # Excel mode: no validation needed (uses default path)
            return True

        elif self._screen == self.SCREEN_DATE:
            val = self._date_entry.get().strip()
            if val and (len(val) != 8 or not val.isdigit()):
                return False
            self._data["first_war_date"] = val if val else None
            return True

        return True

    def _show_current_screen(self):
        """Render the current screen."""
        screens = [
            self._show_mode_screen,
            self._show_season_screen,
            self._show_link_or_path_screen,
            self._show_date_screen,
            self._show_roster_screen,
            self._show_auth_screen,
            self._show_done_screen,
        ]

        self._clear_content()
        screens[self._screen](self.content)

        # Update button states
        self.btn_back.configure(state="normal" if self._screen > 0 else "disabled")

        if self._screen == self.SCREEN_DONE:
            self.btn_next.configure(text="Finish")
        else:
            self.btn_next.configure(text="Next →")

    # -------------------------------------------------------------------------
    # Screen Renderers
    # -------------------------------------------------------------------------

    def _show_mode_screen(self, parent):
        """Screen 0: Choose data source mode."""
        ctk.CTkLabel(
            parent,
            text="Select Data Source",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(0, 20))

        self._mode_var = ctk.StringVar(value="google_sheet")

        ctk.CTkRadioButton(
            parent,
            text="Google Sheets (requires auth setup)",
            variable=self._mode_var,
            value="google_sheet",
            font=("Segoe UI", 13),
        ).pack(anchor="w", padx=(20, 0), pady=(10, 5))

        ctk.CTkRadioButton(
            parent,
            text="Excel (local file)",
            variable=self._mode_var,
            value="excel",
            font=("Segoe UI", 13),
        ).pack(anchor="w", padx=(20, 0), pady=(5, 15))

        ctk.CTkLabel(
            parent,
            text="Google Sheets: Paste link to your copied template spreadsheet\n"
                 "Excel: Uses GW Tracker.xlsx in this folder",
            font=("Segoe UI", 11),
            text_color="gray50",
            justify="left",
        ).pack(anchor="w", padx=(20, 0), pady=(10, 0))

    def _show_season_screen(self, parent):
        """Screen 1: Enter season name."""
        ctk.CTkLabel(
            parent,
            text="Season Name",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(0, 20))

        self._season_entry = ctk.CTkEntry(
            parent,
            placeholder_text="2026-1",
            font=("Segoe UI", 14),
            width=200,
            justify="center",
        )
        self._season_entry.pack(pady=(0, 10))
        self._season_entry.focus_set()

        ctk.CTkLabel(
            parent,
            text="Format: YYYY-N (e.g. 2026-1, 2026-2)",
            font=("Segoe UI", 11),
            text_color="gray50",
        ).pack()

    def _show_link_or_path_screen(self, parent):
        """Screen 2: Google Sheets link OR Excel path (based on mode)."""
        mode = self._data.get("mode", "google_sheet")

        if mode == "google_sheet":
            ctk.CTkLabel(
                parent,
                text="Google Sheets Link",
                font=("Segoe UI", 16, "bold"),
            ).pack(pady=(0, 20))

            self._link_entry = ctk.CTkEntry(
                parent,
                placeholder_text="https://docs.google.com/spreadsheets/d/...",
                font=("Segoe UI", 12),
                width=450,
            )
            self._link_entry.pack(pady=(0, 10))

            ctk.CTkLabel(
                parent,
                text="Paste your copied template URL here",
                font=("Segoe UI", 11),
                text_color="gray50",
            ).pack()
        else:
            # Excel mode - show info, no entry needed
            ctk.CTkLabel(
                parent,
                text="Excel Workbook",
                font=("Segoe UI", 16, "bold"),
            ).pack(pady=(0, 20))

            ctk.CTkLabel(
                parent,
                text="The parser will use:\n\n"
                     "GW Tracker.xlsx\n\n"
                     "(must be in the same folder as this script)",
                font=("Segoe UI", 13),
                justify="center",
            ).pack(pady=(20, 0))

    def _show_date_screen(self, parent):
        """Screen 3: Enter first war date (optional)."""
        ctk.CTkLabel(
            parent,
            text="First War Date",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(0, 20))

        self._date_entry = ctk.CTkEntry(
            parent,
            placeholder_text="20260323",
            font=("Segoe UI", 14),
            width=200,
            justify="center",
        )
        self._date_entry.pack(pady=(0, 10))

        ctk.CTkLabel(
            parent,
            text="Format: YYYYMMDD (first war of the season)\n"
                 "Leave blank to enter it in the spreadsheet later",
            font=("Segoe UI", 11),
            text_color="gray50",
            justify="center",
        ).pack()

    def _show_roster_screen(self, parent):
        """Screen 4: Enter roster names (optional)."""
        ctk.CTkLabel(
            parent,
            text="Roster Entry",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(0, 10))

        ctk.CTkLabel(
            parent,
            text="Paste player names, one per line:",
            font=("Segoe UI", 12),
        ).pack(anchor="w", padx=(0, 0), pady=(5, 2))

        self._roster_textbox = ctk.CTkTextbox(
            parent,
            font=("Consolas", 12),
            height=150,
            wrap="char",
        )
        self._roster_textbox.pack(fill="both", expand=True, padx=(0, 0), pady=(5, 0))

        btn_row = ctk.CTkFrame(parent, fg_color="transparent")
        btn_row.pack(fill="x", pady=(10, 0))

        ctk.CTkButton(
            btn_row,
            text="Skip (I'll enter them in the spreadsheet)",
            command=self._skip_roster,
            width=280,
        ).pack(side="left", padx=(0, 5))

        ctk.CTkButton(
            btn_row,
            text="Confirm",
            command=self._confirm_roster,
            width=120,
        ).pack(side="right", padx=(5, 0))

    def _show_auth_screen(self, parent):
        """Screen 5: Google Sheets authentication (GS only)."""
        ctk.CTkLabel(
            parent,
            text="Google Sheets Authentication",
            font=("Segoe UI", 16, "bold"),
        ).pack(pady=(0, 20))

        ctk.CTkLabel(
            parent,
            text="A browser window will open to authorize access.\n"
                 "Complete the login, then click Continue.",
            font=("Segoe UI", 12),
            justify="center",
        ).pack(pady=(20, 10))

    def _show_done_screen(self, parent):
        """Screen 6: Setup complete."""
        ctk.CTkLabel(
            parent,
            text="Setup Complete!",
            font=("Segoe UI", 22, "bold"),
            text_color="#4CAF50",
        ).pack(pady=(50, 0))

        ctk.CTkLabel(
            parent,
            text="You can now record wars from the main menu.",
            font=("Segoe UI", 13),
        ).pack(pady=(10, 0))

    # -------------------------------------------------------------------------
    # Roster Actions
    # -------------------------------------------------------------------------

    def _skip_roster(self):
        """Skip roster entry and proceed."""
        self._data["roster_lines"] = []
        self._go_next()

    def _confirm_roster(self):
        """Confirm roster entry and proceed."""
        text = self._roster_textbox.get("1.0", "end").strip()
        self._data["roster_lines"] = [
            line.strip() for line in text.splitlines() if line.strip()
        ]
        self._go_next()

    # -------------------------------------------------------------------------
    # Completion
    # -------------------------------------------------------------------------

    def _finish(self):
        """Build the final config dict and close the wizard."""
        mode = self._data.get("mode", "google_sheet")
        season = self._data.get("season_name", "2026-1")

        config = {
            "mode": mode,
            "override_col": "C",
            "date_col": 15,
            "guild_name": None,

        }

        if mode == "google_sheet":
            config["google_sheet"] = {
                "spreadsheet_id": self._data["spreadsheet_id"],
                "source_sheet": season,
                "target_sheet": season,
            }
        else:
            config["excel"] = {
                "workbook_path": "GW Tracker.xlsx",
                "sheet_name": season,
            }

        if self._data.get("first_war_date"):
            config["first_war_date"] = self._data["first_war_date"]

        if self._data.get("roster_lines"):
            config["roster_lines"] = self._data["roster_lines"]

        self.result = config
        self.grab_release()
        self.destroy()

    def run(self):
        """Shows the wizard, blocks until user completes or cancels."""
        self.grab_set()
        self.wait_window()
        return self.result


# ============================================================================
# PARSER IMPORT
# ============================================================================

MAIN_PARSER = None


def import_parser_module(mode: str):
    """Import the appropriate parser module based on mode."""
    if mode == "excel":
        module_name = "gw_parser_excel_v1.py"
    else:
        module_name = "gw_parser_gs_v1.py"

    module_path = BASE_DIR / module_name
    if not module_path.exists():
        raise SystemExit(f"Cannot find: {module_name}")

    spec = importlib.util.spec_from_file_location("parser_module", str(module_path))
    mod = importlib.util.module_from_spec(spec)
    sys.modules["parser_module"] = mod
    spec.loader.exec_module(mod)
    return mod


def get_parser(mode: str = "google_sheet"):
    """Get or load the parser module."""
    global MAIN_PARSER
    if MAIN_PARSER is None:
        MAIN_PARSER = import_parser_module(mode)
    return MAIN_PARSER


# ============================================================================
# MAIN APPLICATION
# ============================================================================

class GwParserApp(ctk.CTk):
    """Main GUI application for the Guild War Parser."""

    def __init__(self):
        # Check for first run BEFORE initializing the window
        self._needs_wizard = not CONFIG_PATH.exists()
        self._wizard_result = None

        super().__init__()
        self.title("Guild War Parser")
        self.geometry("900x660")
        self.protocol("WM_DELETE_WINDOW", self._on_close)

        # Stdio redirectors
        self.stdout_redirect = StdoutRedirector()
        self.stdin_redirect = StdinRedirector()

        # State
        self.running = False
        self.temp_dir = None
        self.stop_event = threading.Event()
        self._aliases_working = None
        self._current_screenshots = None

        # Build UI
        self._build_ui()
        self._start_flush_thread()

        # Launch wizard if needed (after UI is ready)
        if self._needs_wizard:
            self.after(100, self._launch_wizard)

    # -------------------------------------------------------------------------
    # UI Construction
    # -------------------------------------------------------------------------

    def _build_ui(self):
        """Build the main application UI."""
         # Title bar
        title = ctk.CTkFrame(self, height=50)
        title.pack(fill="x")
        
        # Load mode for title
        mode_text = "Google Sheets"
        if CONFIG_PATH.exists():
            try:
                cfg = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
                if cfg.get("mode") == "excel":
                    mode_text = "Excel"
            except Exception:
                pass
        
        ctk.CTkLabel(
            title,
            text=f"  Guild War Parser [{mode_text}]",
            font=("Segoe UI", 18, "bold"),
        ).pack(side="left", padx=5)
        
        # Main container
        container = ctk.CTkFrame(self)
        container.pack(fill="both", expand=True, padx=10, pady=10)

        # Sidebar
        sidebar = ctk.CTkFrame(container, width=200)
        sidebar.pack(side="left", fill="y", padx=(0, 10))
        sidebar.pack_propagate(False)

        ctk.CTkLabel(
            sidebar,
            text="Actions",
            font=("Segoe UI", 13, "bold"),
        ).pack(pady=(10, 20))

        # Action buttons
        buttons = [
            ("Record War", self._on_record_war),
            ("Sync Roster", self._on_sync_roster),
            ("New Season", self._on_new_season),
            ("Export Record", self._on_export),
            ("Configure", self._on_configure),
        ]
        for text, cmd in buttons:
            ctk.CTkButton(
                sidebar,
                text=text,
                command=cmd,
                height=38,
            ).pack(fill="x", padx=12, pady=3)

        # Spacer
        ctk.CTkLabel(sidebar, text="").pack(expand=True)

        # Exit button
        ctk.CTkButton(
            sidebar,
            text="Exit",
            command=self._on_close,
            height=38,
        ).pack(fill="x", padx=12, pady=(0, 10))

        # Right panel
        right = ctk.CTkFrame(container)
        right.pack(side="right", fill="both", expand=True)

        # Console header
        header = ctk.CTkFrame(right, height=32)
        header.pack(fill="x")
        ctk.CTkLabel(
            header,
            text="  Console",
            font=("Segoe UI", 11, "bold"),
        ).pack(side="left", padx=5)

        # Console output
        self.console = ctk.CTkTextbox(
            right,
            font=("Consolas", 11),
            state="disabled",
            wrap="char",
        )
        self.console.pack(fill="both", expand=True, padx=(5, 0), pady=(0, 5))

        # Input bar (hidden by default)
        self.input_bar = ctk.CTkFrame(right, height=40)
        self.input_bar.pack(fill="x", side="bottom", padx=(5, 0), pady=(0, 5))
        self.input_bar.pack_forget()

        self.input_label = ctk.CTkLabel(
            self.input_bar,
            text=">",
            font=("Consolas", 11),
            width=40,
            anchor="w",
        )
        self.input_label.pack(side="left", padx=(5, 0))

        self.input_field = ctk.CTkEntry(self.input_bar, height=28)
        self.input_field.pack(side="left", fill="x", expand=True, padx=(5, 10), pady=6)
        self.input_field.bind("<Return>", self._on_input_submit)

        # Stop button frame (hidden by default)
        self.stop_frame = ctk.CTkFrame(right, height=50)
        self.stop_frame.pack(fill="x", side="bottom", padx=(5, 0), pady=(0, 5))
        self.stop_frame.pack_forget()

        self.stop_btn = ctk.CTkButton(
            self.stop_frame,
            text="Stop Scrolling - Click when done",
            command=self._on_stop_capture,
            height=36,
        )
        self.stop_btn.pack(fill="x", padx=10, pady=6)

        # Status bar (hidden by default)
        self.status_bar = ctk.CTkFrame(right, height=28)
        self.status_bar.pack(fill="x", side="bottom", padx=(5, 0), pady=(0, 5))
        self.status_bar.pack_forget()

        self.status_label = ctk.CTkLabel(
            self.status_bar,
            text="Ready",
            font=("Segoe UI", 10),
        )
        self.status_label.pack(side="left", padx=5)

        # Show status bar initially
        self.status_bar.pack(fill="x", side="bottom", padx=(5, 0), pady=(0, 5))

    # -------------------------------------------------------------------------
    # Console Management
    # -------------------------------------------------------------------------

    def _flush_loop(self):
        """Background thread: flush stdout queue to console."""
        while True:
            try:
                text = self.stdout_redirect.queue.get(timeout=0.1)
                self.console.configure(state="normal")
                self.console.insert("end", text)
                self.console.see(ctk.END)
                self.console.configure(state="disabled")
            except queue.Empty:
                pass
            except Exception:
                break

    def _start_flush_thread(self):
        """Start the console flush thread."""
        threading.Thread(target=self._flush_loop, daemon=True).start()

    def _clear(self):
        """Clear the console output."""
        self.console.configure(state="normal")
        self.console.delete("1.0", "end")
        self.console.configure(state="disabled")

    def _log(self, text=""):
        """Write text to the console via stdout redirector."""
        if text:
            self.stdout_redirect.write(text + "\n")

    # -------------------------------------------------------------------------
    # Status Bar Management
    # -------------------------------------------------------------------------

    def _set_status(self, text):
        """Show status bar with text, hide input/stop."""
        self.input_bar.pack_forget()
        self.stop_frame.pack_forget()
        self.status_bar.pack(fill="x", side="bottom", padx=(5, 0), pady=(0, 5))
        self.status_label.configure(text=text)

    def _show_input(self):
        """Show input bar, hide status/stop."""
        self.status_bar.pack_forget()
        self.stop_frame.pack_forget()
        self.input_bar.pack(fill="x", side="bottom", padx=(5, 0), pady=(0, 5))
        self.input_field.delete(0, "end")
        self.input_field.focus_set()

    def _hide_input(self):
        """Hide input bar."""
        self.input_bar.pack_forget()

    def _show_stop_button(self):
        """Show stop button, hide input/status."""
        self.input_bar.pack_forget()
        self.status_bar.pack_forget()
        self.stop_frame.pack(fill="x", side="bottom", padx=(5, 0), pady=(0, 5))

    def _hide_stop_button(self):
        """Hide stop button."""
        self.stop_frame.pack_forget()

    # -------------------------------------------------------------------------
    # Event Handlers
    # -------------------------------------------------------------------------

    def _on_input_submit(self):
        """Handle Enter key in input field."""
        text = self.input_field.get().strip()
        if text:
            self.stdin_redirect.queue.put(text)

    def _on_stop_capture(self):
        """Handle stop capture button click."""
        self.stop_event.set()

    def _on_record_war(self):
        """Handle Record War button click."""
        if self.running:
            self._log("Already running.")
            return
        
        # Step 1: Choose source
        source_dialog = SourceChoiceDialog(self)
        self.wait_window(source_dialog)
        if source_dialog.result is None:
            return
        source_mode = source_dialog.result
        
        # Step 2: Get date/folder name
        date_dialog = DateDialog(self)
        self.wait_window(date_dialog)
        if date_dialog.result is None:
            return
        
        self._clear()
        self._run_in_thread(self._record_war, args=(date_dialog.result, source_mode))

    def _on_new_season(self):
        """Handle New Season button click."""
        if self.running:
            self._log("Already running.")
            return
        
        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        if config.get("mode", "google_sheet") == "google_sheet":
            current = config["google_sheet"]["target_sheet"]
        else:
            current = config["excel"]["sheet_name"]
        
        dialog = NewSeasonDialog(self, current)
        self.wait_window(dialog)
        if dialog.result is None:
            return
        
        self._clear()
        self._log("Creating new season...\n")
        args = [
            sys.executable, str(BASE_DIR / "sync_roster.py"), "--new-season",
            "--season-name", dialog.result["season_name"] or "",
            "--first-date", dialog.result["first_date"],
        ]
        self._run_sub(args)

    def _on_sync_roster(self):
        """Handle Sync Roster button click."""
        if self.running:
            self._log("Already running.")
            return
        self._clear()
        self._run_sub([sys.executable, str(BASE_DIR / "sync_roster.py")])

    def _on_configure(self):
        """Handle Configure button click."""
        if self.running:
            self._log("Already running.")
            return
        
        dialog = ConfigureDialog(self)
        self.wait_window(dialog)
        if dialog.result is None:
            return
        
        self._clear()
        if dialog.result == "2":
            self._launch_wizard()
        elif dialog.result == "1":
            config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
            mode = config.get("mode", "google_sheet")
            
            if mode == "google_sheet":
                current_id = config["google_sheet"].get("spreadsheet_id", "")
                url_dialog = ChangeSpreadsheetDialog(self, current_id)
                self.wait_window(url_dialog)
                if url_dialog.result is None:
                    self._log("Cancelled.")
                    return
                self._run_sub([
                    sys.executable,
                    str(BASE_DIR / "first_run_setup.py"),
                    "--set-spreadsheet-id", url_dialog.result,
                ])
            else:
                self._log("Excel mode: Change workbook path in config.json manually,")
                self._log("or re-run the full setup wizard.")

    def _on_export(self):
        """Handle Export Record button click."""
        if self.running:
            self._log("Already running.")
            return
        dialog = DateDialog(self)
        self.wait_window(dialog)
        if dialog.result is None:
            return
        self._clear()
        self._log(f"Exporting: {dialog.result}\n")
        self._run_sub([
            sys.executable,
            str(BASE_DIR / "war_record_exporter.py"),
            dialog.result,
        ])

    def _on_close(self):
        """Handle window close."""
        self.stop_event.set()
        self.quit()

    # -------------------------------------------------------------------------
    # Thread Management
    # -------------------------------------------------------------------------

    def _run_in_thread(self, target, args=(), kwargs=None):
        """Run a function in a background thread with redirected stdio."""
        def wrapper():
            old_out = sys.stdout
            old_in = sys.stdin
            sys.stdout = self.stdout_redirect
            sys.stdin = self.stdin_redirect
            try:
                target(*args, **(kwargs or {}))
            except SystemExit:
                pass
            except Exception as exc:
                self._log(f"\nError: {exc}")
            finally:
                sys.stdout = old_out
                sys.stdin = old_in
                self.running = False
                self._set_status("Ready")

        self.running = True
        threading.Thread(target=wrapper, daemon=True).start()

    def _run_sub(self, args):
        """Run a subprocess and stream output to console."""
        def wrapper():
            self._set_status("Running...")
            try:
                proc = subprocess.Popen(
                    args,
                    stdin=subprocess.PIPE,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    cwd=str(BASE_DIR),
                )
                while True:
                    line = proc.stdout.readline()
                    if not line:
                        break
                    self._log(line.rstrip("\n"))
                proc.wait()
                if proc.returncode != 0:
                    self._log(f"\nExited with code {proc.returncode}")
            except FileNotFoundError:
                self._log(f"\nError: Script not found: {args[1]}")
            except Exception as exc:
                self._log(f"\nError: {exc}")
            finally:
                self.running = False
                self._set_status("Ready")

        self.running = True
        threading.Thread(target=wrapper, daemon=True).start()
        
    # -------------------------------------------------------------------------
    # Recording
    # -------------------------------------------------------------------------

    def _record_war(self, folder_name, source_mode):
        """Execute the war recording flow."""
        # Load config to determine mode
        config = json.loads(CONFIG_PATH.read_text(encoding="utf-8"))
        mode = config.get("mode", "google_sheet")
        p = get_parser(mode)

        old_out = sys.stdout
        old_in = sys.stdin
        sys.stdout = self.stdout_redirect
        sys.stdin = self.stdin_redirect

        try:
            if source_mode == "folder":
                screenshots_folder = SCREENSHOTS_ROOT / folder_name
                if not screenshots_folder.exists():
                    raise FileNotFoundError(f"Screenshot folder not found: {screenshots_folder}")
                self._log(f"Data source: Screenshot folder")
                self._log(f"Folder: {screenshots_folder}\n")
                self._current_screenshots = screenshots_folder
            else:
                self._log("Data source: Live Screen Capture")
                self.temp_dir = Path(tempfile.mkdtemp(prefix="gw_capture_"))

                self._log("\nSwitch to the game and open Guild War logs.")
                for i in range(10, 0, -1):
                    self._set_status(f"Recording starts in {i}s...")
                    time.sleep(1)

                self._set_status("Recording - click Stop when done")
                self._show_stop_button()
                

                self.stop_event.clear()
                capture_thread = threading.Thread(
                    target=p.live_capture_loop,
                    args=(self.temp_dir, self.stop_event),
                    daemon=True,
                )
                capture_thread.start()
                self.stop_event.wait()
                self._hide_stop_button()
                self._set_status("Processing OCR...")
                screenshots_folder = self.temp_dir
                self._current_screenshots = screenshots_folder
                self._log("Capture complete.\n")

            if mode == "google_sheet":
                self._record_war_google_sheets(p, config, screenshots_folder, folder_name)
            else:
                self._record_war_excel(p, screenshots_folder, folder_name, config)

        finally:
            sys.stdout = old_out
            sys.stdin = old_in
            if self.temp_dir and self.temp_dir.exists():
                shutil.rmtree(self.temp_dir)
                self._log("Cleaned up capture files.")
            self._set_status("Ready")
            
            
    def _record_war_google_sheets(self, p, config, screenshots_folder, folder_name):
        """Recording flow for Google Sheets mode."""
        client = p.get_gspread_client()
        spreadsheet = client.open_by_key(config["google_sheet"]["spreadsheet_id"])
        source_ws = spreadsheet.worksheet(config["google_sheet"]["source_sheet"])
        target_ws = spreadsheet.worksheet(config["google_sheet"]["target_sheet"])

        source_data = source_ws.get_all_values()
        target_data = target_ws.get_all_values()

        self._process_war_data(p, source_data, target_data, target_ws, folder_name, config)

    def _record_war_excel(self, p, screenshots_folder, folder_name, config):
        """Recording flow for Excel mode."""
        from openpyxl import load_workbook

        wb_path = BASE_DIR / config["excel"]["workbook_path"]
        lock_file = wb_path.parent / f"~${wb_path.name}"
        if lock_file.exists():
            raise SystemExit(
                f"Workbook is open in Excel. Close it and try again.")

        if not wb_path.exists():
            raise FileNotFoundError(f"Workbook not found: {wb_path}")

        wb = load_workbook(wb_path)
        wb_values = load_workbook(wb_path, data_only=True, read_only=True)

        sheet_name = config["excel"]["sheet_name"]
        if sheet_name not in wb.sheetnames:
            raise RuntimeError(f"Sheet '{sheet_name}' not found in workbook.")

        source_ws = wb[sheet_name]
        target_ws = wb[sheet_name]
        target_ws_values = wb_values[sheet_name]

        # Convert to 2D lists for uniform processing
        source_data = self._excel_to_2d_list(source_ws)
        target_data = self._excel_to_2d_list(target_ws_values)

        self._process_war_data(p, source_data, target_data, target_ws, folder_name, config)

        # Save workbook
        wb.save(wb_path)
        self._log(f"\nWorkbook saved: {wb_path.name}")

    def _excel_to_2d_list(self, ws, max_col: int = 650) -> list:
        """Convert openpyxl worksheet to 2D list."""
        data = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
            data.append([str(v) if v is not None else "" for v in row])
        return data

    def _process_war_data(self, p, source_data, target_data, target_ws, folder_name, config):
        """Process war data, show results, ask for confirmation before writing."""
        roster = p.read_roster(source_data)
        overrides = p.read_column_a_overrides(
            target_data, roster, config.get("override_col", "C"),
        )

        aliases_existing = p.load_aliases(p.ALIASES_PATH)
        aliases_working = dict(aliases_existing)
        self._aliases_working = aliases_working
        aliases_created = {}

        self._log(f"Roster: {len(roster)} players\n")

        extracted, _ = p.extract_rows_from_folder(
            self._current_screenshots, folder_name, roster,
            aliases_existing, aliases_working, aliases_created, overrides,
        )
        normalized = p.build_normalized_table(roster, extracted)

        date_col = config.get("date_col", 15)
        target_row = p.find_date_row(target_data, folder_name, date_col)
        target_maps = p.build_target_section_maps(target_data)
        prev_active = p.get_previous_war_active_status(
            target_data, target_row, target_maps, date_col,
        )

        # Interactive prompts for edge cases
        self._handle_edge_cases(normalized, extracted, prev_active)

        p.print_results_output_terminal(extracted)
        p.print_normalized_table_terminal(normalized)
        p.print_run_summary(extracted, normalized, aliases_created)

        if len(extracted) < p.MIN_EXPECTED_MATCHED_PLAYERS:
            self._log(
                f"\nWarning: {len(extracted)} matched "
                f"(threshold: {p.MIN_EXPECTED_MATCHED_PLAYERS}). Continue? [y/N]: "
            )
            self._show_input()
            ans = self.stdin_redirect.queue.get().strip().lower()
            self._hide_input()
            if ans not in ("y", "yes"):
                self._log("Cancelled.")
                return

        # Confirmation before writing
        self._log("\nWrite results to spreadsheet? [y/N]: ")
        self._show_input()
        ans = self.stdin_redirect.queue.get().strip().lower()
        self._hide_input()
        if ans not in ("y", "yes"):
            self._log("Cancelled. No changes made.")
            return

        # Write to workbook
        self._write_war_data(p, target_data, target_ws, target_row, normalized, target_maps, aliases_created, folder_name, config)
    
        
    def _write_war_data(self, p, target_data, target_ws, target_row, normalized, target_maps, aliases_created, folder_name, config):
        """Write processed war data to spreadsheet and export PNG."""
        mode = config.get("mode", "google_sheet")

        has_data = p.row_has_existing_values(target_data, target_row, target_maps)
        if has_data:
            self._log("Overwrite existing data? [y/N]: ")
            self._show_input()
            ans = self.stdin_redirect.queue.get().strip().lower()
            self._hide_input()
            if ans not in ("y", "yes"):
                self._log("Cancelled.")
                return

        if has_data:
            p.clear_target_row_values(target_ws, target_row, target_maps)

        p.write_to_rough_sheet(target_ws, target_row, normalized, target_maps)

        if aliases_created:
            p.save_aliases(p.ALIASES_PATH, self._aliases_working)

        if mode == "google_sheet":
            self._log(f"Sheet: {config['google_sheet']['target_sheet']}")
        else:
            self._log(f"Sheet: {config['excel']['sheet_name']}")

        self._log(f"Aliases saved: {len(aliases_created)}")

        # Export PNG
        self._log("\nExporting war record PNG...")
        try:
            if mode == "google_sheet":
                from war_record_exporter import export_to_png
                export_to_png(
                    p.get_gspread_client(),
                    config["google_sheet"]["spreadsheet_id"],
                    config["google_sheet"]["target_sheet"],
                    folder_name,
                )
            else:
                from war_record_exporter import export_to_png_excel
                export_to_png_excel(folder_name, config)
        except Exception as exc:
            self._log(f"Warning: Export failed: {exc}")

        self._log("\nDone.")
        

    def _handle_edge_cases(self, normalized, extracted, prev_active):
        """Handle interactive prompts for new recruits and missing players."""
        for row in normalized:
            name = row["Name"]

            # New recruit with 0 tokens
            if (
                row["Active flag"] == 1
                and prev_active is not None
                and not prev_active.get(row["Name"], False)
                and row["Tokens Used"] == 0
            ):
                self._log(f"\n{name} - new recruit, 0 tokens. Participated? [y/N]: ")
                self._show_input()
                ans = self.stdin_redirect.queue.get().strip().lower()
                self._hide_input()
                if ans in ("y", "yes"):
                    row["Notes"] = "Active (New recruit, confirmed 0 stats)"
                else:
                    self._log(f"  -> Marking {name} as inactive.")
                    row.update({
                        "Active flag": "", "Tokens Used": "",
                        "Offense W": "", "Offense D": "", "Offense L": "",
                        "Defense W": "", "Defense D": "", "Defense L": "",
                        "Notes": "Inactive (New recruit)",
                    })
                    extracted = [r for r in extracted if r.matched_name != row["Name"]]
                continue

            # Missing from screenshots
            if row["Notes"] == "Inactive (not found in screenshots)":
                should_ask = (
                    prev_active is None
                    or prev_active.get(row["Name"], False)
                )
                if should_ask:
                    self._log(f"\n{name} was active, missing now.")
                    self._log("  1. Manually enter stats")
                    self._log("  2. Mark as Inactive")
                    self._log("  3. Stop run")
                    self._show_input()
                    ans = self.stdin_redirect.queue.get().strip()
                    self._hide_input()
                    if ans == "1":
                        # Offense with validation
                        ow, od, ol = 0, 0, 0
                        while True:
                            self._log("  Offense W D L (e.g. 2 0 1): ")
                            self._show_input()
                            off_input = self.stdin_redirect.queue.get().strip()
                            self._hide_input()
                            off = off_input.split()
                            if len(off) == 3 and all(p.isdigit() for p in off):
                                ow, od, ol = int(off[0]), int(off[1]), int(off[2])
                                break
                            self._log("  Invalid format. Enter 3 numbers separated by spaces.")
        
                        # Defense with validation
                        dw, dd, dl = 0, 0, 0
                        while True:
                            self._log("  Defense W D L (e.g. 2 0 1): ")
                            self._show_input()
                            def_input = self.stdin_redirect.queue.get().strip()
                            self._hide_input()
                            de = def_input.split()
                            if len(de) == 3 and all(p.isdigit() for p in de):
                                dw, dd, dl = int(de[0]), int(de[1]), int(de[2])
                                break
                            self._log("  Invalid format. Enter 3 numbers separated by spaces.")
        
                        row.update({
                            "Active flag": 1,
                            "Tokens Used": ow + od + ol,
                            "Offense W": ow, "Offense D": od, "Offense L": ol,
                            "Defense W": dw, "Defense D": dd, "Defense L": dl,
                            "Notes": "Active (Manual)",
                        })
                    elif ans != "2":
                        self._log("Stopped by user.")
                        raise RuntimeError("Stopped by user")

    # -------------------------------------------------------------------------
    # Wizard
    # -------------------------------------------------------------------------

    def _launch_wizard(self):
        """Launch the first-run wizard and call setup script."""
        wizard = FirstRunWizard(self)
        result = wizard.run()

        if result is None:
            self._log("Setup cancelled. You can re-run from Configure button.")
            return

        mode = result.get("mode", "google_sheet")
        season = result.get("season_name", "2026-1")

        # Build args for first_run_setup.py
        args = [
            sys.executable,
            str(BASE_DIR / "first_run_setup.py"),
            "--mode", mode,
            "--season", season,
        ]

        if mode == "google_sheet" and result.get("spreadsheet_id"):
            args.extend(["--spreadsheet-id", result["spreadsheet_id"]])

        if result.get("first_war_date"):
            args.extend(["--first-date", result["first_war_date"]])

        if result.get("roster_lines"):
            args.extend(result["roster_lines"])

        # Google Sheets: run auth first
        if mode == "google_sheet":
            self._log("Running authentication...\n")
            self._run_sub([
                sys.executable,
                str(BASE_DIR / "setup_auth.py"),
            ])
            while self.running:
                time.sleep(0.1)

        # Run setup script
        self._log("Running setup...\n")
        self._run_sub(args)


# ============================================================================
# ENTRY POINT
# ============================================================================

if __name__ == "__main__":
    app = GwParserApp()
    app.mainloop()