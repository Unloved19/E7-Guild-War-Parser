"""Syncs Column B roster names with helper section columns, or creates a new season tab.

Run standalone:
    python sync_roster.py              # Sync add/remove
    python sync_roster.py --new-season # Duplicate tab, clear data, keep roster+formulas
"""

import sys
import json
from pathlib import Path
from typing import Dict, List, Optional, Set

import gspread
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURATION
# ============================================================================

BASE_DIR = Path(__file__).resolve().parent
CONFIG_PATH = BASE_DIR / "config.json"
TOKEN_PATH = BASE_DIR / "token.json"

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive.readonly",
]

# All 10 sections that have 60 player columns.
# MVP and MVP Count are NOT included — they have no player helper columns.
WRITE_SECTIONS = [
    "Roster Active Status", "Tokens Tracking",
    "Offense Wins", "Offense Draws", "Offense Losses", "Offense Win Rate",
    "Defense Wins", "Defense Draws", "Defense Losses", "Defense Win Rate",
]

SECTION_HEADERS = [
    "Roster Active Status", "Tokens Tracking",
    "Offense Wins", "Offense Draws", "Offense Losses", "Offense Win Rate",
    "MVP", "MVP Count",
    "Defense Wins", "Defense Draws", "Defense Losses", "Defense Win Rate",
]

# Sections that hold DATA (not formulas) - used for new season clear
NEW_SEASON_DATA_SECTIONS = [
    "Roster Active Status", "Tokens Tracking",
    "Offense Wins", "Offense Draws", "Offense Losses",
    "Defense Wins", "Defense Draws", "Defense Losses",
]

def load_config() -> dict:
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def save_config(config: dict) -> None:
    CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")


def get_gspread_client() -> gspread.Client:
    creds = Credentials.from_authorized_user_file(str(TOKEN_PATH), SCOPES)
    if not creds.valid and creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return gspread.authorize(creds)

# ============================================================================
# SHARED DATA ACCESS (Works with 2D lists from both backends)
# ============================================================================

def cell_val(data: list, row: int, col: int):
    if row < 1 or row > len(data):
        return None
    r = data[row - 1]
    if col < 1 or col > len(r):
        return None
    v = r[col - 1]
    return v if v != "" else None


def read_roster(data: list) -> List[str]:
    roster = []
    blank = 0
    row = 5
    while row <= len(data) + 10:
        val = cell_val(data, row, 2)
        if val is None:
            blank += 1
            if blank >= 10:
                break
        else:
            blank = 0
            name = str(val).strip()
            if name.lower().startswith("notes:"):
                row += 1
                continue
            roster.append(name)
        row += 1
    return roster


def find_section_starts(data: list) -> Dict[str, int]:
    starts = {}
    max_col = max((len(r) for r in data), default=0)
    for col in range(1, max_col + 1):
        val = cell_val(data, 2, col)
        if not val:
            continue
        text = str(val).strip()
        if text in SECTION_HEADERS and text not in starts:
            starts[text] = col
    return starts


def build_section_maps(data: list) -> Dict[str, Dict[str, int]]:
    starts = find_section_starts(data)
    maps = {}
    sorted_h = sorted(starts.items(), key=lambda x: x[1])
    max_col = max((len(r) for r in data), default=0)

    for idx, (header, start_col) in enumerate(sorted_h):
        next_start = sorted_h[idx + 1][1] if idx + 1 < len(sorted_h) else max_col + 1
        name_map = {}
        for col in range(start_col + 1, next_start):
            val = cell_val(data, 2, col)
            if not val:
                continue
            name = str(val).strip()
            if name in SECTION_HEADERS:
                break
            if name != "":
                name_map[name] = col
        maps[header] = name_map
    return maps

# ============================================================================
# COLUMN UTILITIES
# ============================================================================

def _col_letter(col: int) -> str:
    result = ""
    while col > 0:
        col, remainder = divmod(col - 1, 26)
        result = chr(65 + remainder) + result
    return result

# ============================================================================
# EXCEL HELPERS
# ============================================================================

def _excel_to_2d_list(ws, max_col: int = 650) -> list:
    """Converts an openpyxl worksheet to a 2D list matching gspread's format."""
    data = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=max_col, values_only=True):
        data.append([str(v) if v is not None else "" for v in row])
    return data


def _excel_cell_val(data: list, row: int, col: int):
    """Same as cell_val but for the 2D list from openpyxl."""
    if row < 1 or row > len(data):
        return None
    row_data = data[row - 1]
    if col < 1 or col > len(row_data):
        return None
    val = row_data[col - 1]
    return val if val != "" else None

def sheet_date_to_folder_date_str(value) -> Optional[str]:
    from datetime import date, datetime
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date().strftime("%Y%m%d")
    if isinstance(value, date):
        return value.strftime("%Y%m%d")
    if isinstance(value, str):
        value = value.strip()
        for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d", "%d-%m-%Y", "%d-%m-%y", "%Y/%m/%d", "%Y%m%d", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(value, fmt).strftime("%Y%m%d")
            except ValueError:
                continue
    return None


# ============================================================================
# SYNC ROSTER - EXCEL MODE
# ============================================================================

def sync_roster_excel(ws, data: list, section_maps: Dict[str, Dict[str, int]], roster: List[str]) -> None:
    """Excel version of sync_roster. Identical logic, writes via openpyxl."""
    roster_set = set(roster)
    existing: Set[str] = set()
    for section in WRITE_SECTIONS:
        existing.update(section_maps[section].keys())

    to_add = sorted(roster_set - existing)
    to_remove = sorted(existing - roster_set)

    if not to_add and not to_remove:
        print("Roster is already in sync with helper sections.")
        return

    if to_remove:
        print(f"\nClearing {len(to_remove)} removed player(s):")
        for name in to_remove:
            print(f"  - {name}")
        for section in WRITE_SECTIONS:
            for name in to_remove:
                col = section_maps[section].get(name)
                if col:
                    letter = get_column_letter(col)
                    ws.cell(2, col).value = None
                    for row in range(5, ws.max_row + 1):
                        if ws.cell(row, col).value is not None:
                            ws.cell(row, col).value = None

    if to_add:
        print(f"\nAdding {len(to_add)} new player(s):")
        for name in to_add:
            print(f"  + {name}")
        section_starts = find_section_starts(data)
        sorted_h = sorted(section_starts.items(), key=lambda x: x[1])
        max_col = 650 # Same as _excel_to_2d_list max_col to avoid issues with very wide sheets
        for section in WRITE_SECTIONS:
            if section not in section_starts:
                continue
            existing_names = set(section_maps[section].keys())
            start_col = section_starts[section]
            idx = [i for i, h in enumerate(sorted_h) if h[0] == section][0]
            next_start = sorted_h[idx + 1][1] if idx + 1 < len(sorted_h) else max_col + 1
            empty_slots = []
            for col in range(start_col + 1, next_start):
                header = cell_val(data, 2, col)
                if (not header) or (header not in existing_names and header not in SECTION_HEADERS):
                    empty_slots.append(col)
            if len(empty_slots) < len(to_add):
                print(f"  WARNING: '{section}' only has {len(empty_slots)} empty slot(s), "
                      f"need {len(to_add)}!")
            for i, name in enumerate(to_add):
                if i < len(empty_slots):
                    ws.cell(2, empty_slots[i], name)

    print("Done.")

# ============================================================================
# SYNC ROSTER - GOOGLE SHEETS MODE
# ============================================================================

def sync_roster(ws, data: list, section_maps: Dict[str, Dict[str, int]], roster: List[str]) -> None:
    """Adds new players to helper sections, clears removed players.

    Each section is independent — finds empty slots per section.
    No cross-section alignment needed because formulas reference
    their own section ranges.
    """
    roster_set = set(roster)
    existing: Set[str] = set()
    for section in WRITE_SECTIONS:
        existing.update(section_maps[section].keys())

    to_add = sorted(roster_set - existing)
    to_remove = sorted(existing - roster_set)

    if not to_add and not to_remove:
        print("Roster is already in sync with helper sections.")
        return

    updates = []

    # --- Remove players (clear header + data in their column) ---
    if to_remove:
        print(f"\nClearing {len(to_remove)} removed player(s):")
        for name in to_remove:
            print(f"  - {name}")
        for section in WRITE_SECTIONS:
            for name in to_remove:
                col = section_maps[section].get(name)
                if col:
                    letter = _col_letter(col)
                    updates.append({"range": f"{letter}2", "values": [[""]]})
                    for row in range(5, len(data) + 1):
                        if cell_val(data, row, col):
                            updates.append({"range": f"{letter}{row}", "values": [[""]]})

    # --- Add players (find empty slots independently per section) ---
    if to_add:
        print(f"\nAdding {len(to_add)} new player(s):")
        for name in to_add:
            print(f"  + {name}")
        section_starts = find_section_starts(data)
        sorted_h = sorted(section_starts.items(), key=lambda x: x[1])
        max_col = 650 # Same as _excel_to_2d_list max_col to avoid issues with very wide sheets
        for section in WRITE_SECTIONS:
            if section not in section_starts:
                continue
            existing_names = set(section_maps[section].keys())
            start_col = section_starts[section]
            idx = [i for i, h in enumerate(sorted_h) if h[0] == section][0]
            next_start = sorted_h[idx + 1][1] if idx + 1 < len(sorted_h) else max_col + 1
            empty_slots = []
            for col in range(start_col + 1, next_start):
                header = cell_val(data, 2, col)
                if (not header) or (header not in existing_names and header not in SECTION_HEADERS):
                    empty_slots.append(col)
            if len(empty_slots) < len(to_add):
                print(f"  WARNING: '{section}' only has {len(empty_slots)} empty slot(s), "
                      f"need {len(to_add)}!")
            for i, name in enumerate(to_add):
                if i < len(empty_slots):
                    col = empty_slots[i]
                    updates.append({"range": f"{_col_letter(col)}2", "values": [[name]]})

    if updates:
        ws.batch_update(updates)
        total = len(to_add) * len(WRITE_SECTIONS) + len(updates) - (len(to_add) * len(WRITE_SECTIONS))
        print(f"\n  Updated {len(updates)} cells.")
    print("Done.")

# ============================================================================
# NEW SEASON - EXCEL MODE
# ============================================================================

def new_season_excel(wb, config: dict, season_name: str = None, first_date: str = None) -> None:
    """Excel version of new_season. Clears war data only, runs sync."""
    current_name = config["excel"]["sheet_name"]
    
    if season_name:
        new_name = season_name
    else:
        parts = current_name.rsplit("-", 1)
        if len(parts) == 2 and parts[1].strip().isdigit():
            default_new = f"{parts[0]}-{int(parts[1].strip()) + 1}"
        else:
            default_new = f"{current_name} (2)"
        print(f"\nCurrent season sheet: {current_name}")
        new_name = input(f"New season sheet name [default: {default_new}]: ").strip()
        if not new_name:
            new_name = default_new
            
    if not first_date:
        if not season_name:
            # CLI mode - prompt for date
            date_input = input("First war date of new season (YYYYMMDD): ").strip()
            if date_input and len(date_input) == 8 and date_input.isdigit():
                first_date = date_input

    if new_name in wb.sheetnames:
        print(f"Error: Sheet '{new_name}' already exists.")
        return

    source_ws = wb[current_name]

    print(f"\nCreating '{new_name}'...", end=" ", flush=True)
    new_ws = wb.copy_worksheet(source_ws)
    new_ws.title = new_name
    print("Done.")

    # Only clear data cells (rows 5+) in data sections - keep headers in row 2
    print("Clearing war data...", end=" ", flush=True)
    data = _excel_to_2d_list(new_ws)
    section_maps = build_section_maps(data)
    cleared = 0

    for section in NEW_SEASON_DATA_SECTIONS:
        for name, col in section_maps[section].items():
            for row in range(5, new_ws.max_row + 1):
                if new_ws.cell(row, col).value is not None:
                    new_ws.cell(row, col).value = None
                    cleared += 1

    print(f"Cleared {cleared} data cells.")

    # Write first war date to O5
    if first_date:
        from datetime import datetime
        parsed = datetime.strptime(first_date, "%Y%m%d")
        date_col = config.get("date_col", 15)
        new_ws.cell(5, date_col, parsed.date())
        print(f"Wrote first war date to O5: {parsed.strftime('%d/%m/%Y')}")

    # Update config
    config["excel"]["sheet_name"] = new_name
    save_config(config)

    # Run sync roster on new sheet
    print("\nSyncing roster to helper sections...")
    new_data = _excel_to_2d_list(new_ws)
    new_maps = build_section_maps(new_data)
    roster = read_roster(new_data)
    sync_roster_excel(new_ws, new_data, new_maps, roster)

    print(f"\nNew season sheet '{new_name}' is ready.")
    
# ============================================================================
# NEW SEASON - GOOGLE SHEETS MODE
# ============================================================================

def new_season(client, spreadsheet, config: dict, season_name: str = None, first_date: str = None) -> None:
    """Duplicates the current tab, clears war data only, runs sync."""
    current_name = config["google_sheet"]["target_sheet"]
    
    if season_name:
        new_name = season_name
    else:
        parts = current_name.rsplit("-", 1)
        if len(parts) == 2 and parts[1].strip().isdigit():
            default_new = f"{parts[0]}-{int(parts[1].strip()) + 1}"
        else:
            default_new = f"{current_name} (2)"
        print(f"\nCurrent season tab: {current_name}")
        new_name = input(f"New season tab name [default: {default_new}]: ").strip()
        if not new_name:
            new_name = default_new
            
    if not first_date:
        if not season_name:
            # CLI mode - prompt for date
            date_input = input("First war date of new season (YYYYMMDD): ").strip()
            if date_input and len(date_input) == 8 and date_input.isdigit():
                first_date = date_input

    existing_names = [ws.title for ws in spreadsheet.worksheets()]
    if new_name in existing_names:
        print(f"Error: Tab '{new_name}' already exists.")
        return

    source_ws = spreadsheet.worksheet(current_name)

    print(f"\nDuplicating '{current_name}' as '{new_name}'...", end=" ", flush=True)
    body = {"requests": [{"duplicateSheet": {"sourceSheetId": source_ws.id, "insertSheetIndex": 0, "newSheetName": new_name}}]}
    spreadsheet.batch_update(body)
    print("Done.")

    new_ws = spreadsheet.worksheet(new_name)
    data = new_ws.get_all_values()
    section_maps = build_section_maps(data)

    # Only clear data cells (rows 5+) in data sections - keep headers in row 2
    print("Clearing war data...", end=" ", flush=True)
    updates = []
    for section in NEW_SEASON_DATA_SECTIONS:
        for name, col in section_maps[section].items():
            letter = _col_letter(col)
            for row in range(5, len(data) + 1):
                if cell_val(data, row, col):
                    updates.append({"range": f"{letter}{row}", "values": [[""]]})

    if updates:
        new_ws.batch_update(updates)
    print(f"Cleared {len(updates)} data cells.")

    # Write first war date to O5
    if first_date:
        from datetime import datetime
        date_col = config.get("date_col", 15)
        formatted_date = datetime.strptime(first_date, "%Y%m%d").strftime("%d/%m/%Y")
        new_ws.batch_update(
            [{"range": f"{_col_letter(date_col)}5", "values": [[formatted_date]]}],
            value_input_option="USER_ENTERED",
        )
        print(f"Wrote first war date to O5: {formatted_date}")
        
    # Update config
    config["google_sheet"]["source_sheet"] = new_name
    config["google_sheet"]["target_sheet"] = new_name
    save_config(config)

    # Run sync roster on new sheet
    print("\nSyncing roster to helper sections...")
    new_data = new_ws.get_all_values()
    new_maps = build_section_maps(new_data)
    roster = read_roster(new_data)
    sync_roster(new_ws, new_data, new_maps, roster)

    print(f"\nNew season tab '{new_name}' is ready.")

# ============================================================================
# ENTRY POINT & MODE ROUTING
# ============================================================================

def main():
    args = sys.argv[1:]
    
    if "--write-roster" in args:
        config = load_config()
        roster_args = [a for a in args if a != "--write-roster"]
        if not roster_args:
            raise SystemExit("No roster names provided.")
        backend_mode = config.get("mode", "google_sheet")
        if backend_mode == "excel":
            write_roster_to_column_b_excel(roster_args, config)
        else:
            write_roster_to_column_b_gs(roster_args, config)
        return
    
    if "--new-season" in args:
        config = load_config()
        season_name = None
        first_date = None
        
        # Parse optional arguments
        if "--season-name" in args:
            idx = args.index("--season-name")
            season_name = args[idx + 1] if idx + 1 < len(args) else None
        if "--first-date" in args:
            idx = args.index("--first-date")
            first_date = args[idx + 1] if idx + 1 < len(args) else None
        
        backend_mode = config.get("mode", "google_sheet")
        if backend_mode == "excel":
            wb_path = BASE_DIR / config["excel"]["workbook_path"]
            wb = load_workbook(wb_path)
            new_season_excel(wb, config, season_name, first_date)
            wb.save(wb_path)
        else:
            client = get_gspread_client()
            spreadsheet = client.open_by_key(config["google_sheet"]["spreadsheet_id"])
            new_season(client, spreadsheet, config, season_name, first_date)
        return
    
    mode = "sync"
    config = load_config()
    backend_mode = config.get("mode", "google_sheet")
    
    if backend_mode == "excel":
        _main_excel(mode, config)
    else:
        _main_google_sheet(mode, config)


def _main_excel(mode: str, config: dict) -> None:
    wb_path = BASE_DIR / config["excel"]["workbook_path"]
    if not wb_path.exists():
        raise SystemExit(f"Workbook not found: {wb_path}")
    
    if mode == "new-season":
        wb = load_workbook(wb_path)
        new_season_excel(wb, config)
        wb.save(wb_path)
        return
    
    wb = load_workbook(wb_path)
    sheet_name = config["excel"]["sheet_name"]
    
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found in workbook.")
    
    ws = wb[sheet_name]
    print(f"Reading '{sheet_name}'...", end=" ", flush=True)
    data = _excel_to_2d_list(ws)
    print("Done.")
    
    roster = read_roster(data)
    if not roster:
        print("No roster found in Column B. Add player names first.")
        return
    
    print(f"Found {len(roster)} player(s) in Column B.")
    section_maps = build_section_maps(data)
    sync_roster_excel(ws, data, section_maps, roster)
    
    # Write first war date to O5 if empty
    date_col = config.get("date_col", 15)
    if ws.cell(5, date_col).value is None:
        from datetime import date, datetime
        date_val = _excel_cell_val(data, 5, date_col)
        if date_val:
            first_date = sheet_date_to_folder_date_str(date_val)
            if first_date:
                parsed = datetime.strptime(first_date, "%Y%m%d")
                ws.cell(5, date_col, parsed.date())
                print(f"Wrote first war date to O5: {first_date}")
    
    wb.save(wb_path)
    print(f"Workbook saved.")


def _main_google_sheet(mode: str, config: dict) -> None:
    client = get_gspread_client()
    spreadsheet = client.open_by_key(config["google_sheet"]["spreadsheet_id"])
    
    if mode == "new-season":
        new_season(client, spreadsheet, config)
        return
    
    ws = spreadsheet.worksheet(config["google_sheet"]["target_sheet"])
    print(f"Reading '{config['google_sheet']['target_sheet']}'...", end=" ", flush=True)
    data = ws.get_all_values()
    print("Done.")
    
    roster = read_roster(data)
    if not roster:
        print("No roster found in Column B. Add player names first.")
        return
    
    print(f"Found {len(roster)} player(s) in Column B.")
    section_maps = build_section_maps(data)
    sync_roster(ws, data, section_maps, roster)
    
    # Write first war date to O5 if empty
    date_col = config.get("date_col", 15)
    if cell_val(data, 5, date_col) is None:
        date_val = cell_val(data, 5, date_col)
        if date_val:
            first_date = sheet_date_to_folder_date_str(date_val)
            if first_date:
                from datetime import datetime
                parsed = datetime.strptime(first_date, "%Y%m%d")
                updates = [{"range": f"{_col_letter(date_col)}5", "values": [[parsed.strftime("%d/%m/%Y")]]}]
                ws.batch_update(updates)
                print(f"Wrote first war date to O5: {first_date}")
    
    print("Done.")
    
    
# ============================================================================
# WRITE ROSTER TO COLUMN B
# ============================================================================

def write_roster_to_column_b_excel(roster: List[str], config: dict) -> None:
    """Writes roster names to Column B starting at row 5 (Excel mode)."""
    wb_path = BASE_DIR / config["excel"]["workbook_path"]
    
    lock_file = BASE_DIR / f"~${config['excel']['workbook_path']}"
    if lock_file.exists():
        raise SystemExit(
            f"Workbook is currently open in Excel.\n"
            f"Detected lock file: {lock_file.name}\n"
            f"Please save and close the workbook, then try again."
        )
    
    if not wb_path.exists():
        raise SystemExit(f"Workbook not found: {wb_path}")
    
    wb = load_workbook(wb_path)
    sheet_name = config["excel"]["sheet_name"]
    
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Sheet '{sheet_name}' not found in workbook.")
    
    ws = wb[sheet_name]
    
    # Clear existing roster (rows 5 onwards in column B)
    for row in range(5, ws.max_row + 1):
        if ws.cell(row, 2).value is not None:
            ws.cell(row, 2).value = None
    
    # Write new roster
    for idx, name in enumerate(roster, start=5):
        ws.cell(idx, 2, name)
    
    wb.save(wb_path)
    print(f"Wrote {len(roster)} player(s) to Column B in '{sheet_name}'.")


def write_roster_to_column_b_gs(roster: List[str], config: dict) -> None:
    """Writes roster names to Column B starting at row 5 (Google Sheets mode)."""
    client = get_gspread_client()
    spreadsheet = client.open_by_key(config["google_sheet"]["spreadsheet_id"])
    ws = spreadsheet.worksheet(config["google_sheet"]["source_sheet"])
    
    # Read current data to find how many rows to clear
    data = ws.get_all_values()
    max_row = len(data) + 5  # Add buffer
    
    updates = []
    
    # Clear existing roster (rows 5 onwards in column B)
    for row in range(5, max_row + 1):
        if cell_val(data, row, 2) is not None:
            updates.append({"range": f"B{row}", "values": [[""]]})
    
    # Write new roster
    for idx, name in enumerate(roster, start=5):
        updates.append({"range": f"B{idx}", "values": [[name]]})
    
    if updates:
        ws.batch_update(updates)
    
    print(f"Wrote {len(roster)} player(s) to Column B in '{config['google_sheet']['source_sheet']}'.")

if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        print(f"\nError: {exc}")
        raise SystemExit(1)